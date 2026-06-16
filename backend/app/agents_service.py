from __future__ import annotations

import asyncio
import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from functools import lru_cache
import hashlib
import json
import logging
import re
import shutil
import sys
import time
from pathlib import Path
from threading import Event, Lock, Thread
from typing import Any, Callable, Dict, List, Optional

from agents import (
    Agent,
    CodeInterpreterTool,
    ModelSettings,
    ModelRetrySettings,
    RetryDecision,
    RetryPolicyContext,
    Runner,
    SQLiteSession,
    WebSearchTool,
    function_tool,
    set_default_openai_key,
)
from agents.run_context import RunContextWrapper
from agents.mcp import MCPServerStdio, create_static_tool_filter
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .config import get_settings
from .model_builder import (
    fetch_model_builder_state,
    save_custom_calculation_state,
    update_model_builder_state,
    update_model_graph_state,
)
from .model_charts import build_model_node_chart
from .project_memory import (
    compact_project_memory_after_run,
    fetch_project_chat_messages,
    fetch_project_compact_memory,
    persist_project_chat_run,
    search_project_compact_memory,
)
from .storage import ConversationStore
from .validated_variables import (
    apply_transformation_rows,
    compact_validated_data_from_rows,
    list_validated_variable_records,
    run_validated_variable_record,
    save_validated_variable_record,
    validated_data_headers,
    validated_data_latest_rows,
    validated_data_rows,
)


settings = get_settings()
logger = logging.getLogger("abs.backend.agents")
if not logger.handlers:
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(
        logging.Formatter("[%(asctime)s] %(levelname)s %(name)s - %(message)s")
    )
    logger.addHandler(stream_handler)
logger.setLevel(logging.INFO)
logger.propagate = False

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SESSION_DB_PATH = settings.runtime_dir / "agent_sdk_sessions.sqlite3"
SESSION_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
GPT_5_4_INPUT_PRICE_PER_MILLION = 2.50
GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION = 0.625
GPT_5_4_OUTPUT_PRICE_PER_MILLION = 10.00
AI_COST_SURCHARGE_RATE = 0.10
SUPPORTED_OFFICIAL_CUSTOM_DATA_SOURCES = {
    "ABS": ("abs", "abs.gov.au", "australian bureau of statistics", "data.api.abs.gov.au"),
    "OECD": ("oecd", "sdmx.oecd.org"),
    "World Bank": ("world bank", "worldbank", "api.worldbank.org"),
    "IMF": ("imf", "imf.org"),
    "RBA": ("rba", "reserve bank of australia", "rba.gov.au"),
    "UN Comtrade": ("un comtrade", "comtrade", "comtradeapi.un.org"),
}

_CANCELLATION_LOCK = Lock()
_CANCELLATION_EVENTS: Dict[str, Event] = {}


@dataclass
class AgentRuntimeContext:
    conversation_id: str
    store: ConversationStore
    code_container_id: str
    status_callback: Callable[[str], None]
    user_id: str = ""
    project_id: str = ""
    project_name: str = ""
    project_compact_memory: Dict[str, Any] | None = None


class ConversationCancelled(RuntimeError):
    """Raised when a conversation is cancelled mid-generation."""


class NisabaProjectSession:
    """Agents SDK session hydrated from Supabase-backed project chat history."""

    def __init__(self, session_id: str, initial_items: Optional[List[Dict[str, Any]]] = None):
        self.session_id = session_id
        self._items = [copy.deepcopy(item) for item in list(initial_items or [])]
        self._lock = asyncio.Lock()

    async def get_items(self, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        async with self._lock:
            if limit is None:
                return copy.deepcopy(self._items)
            return copy.deepcopy(self._items[-limit:])

    async def add_items(self, items: List[Dict[str, Any]]) -> None:
        async with self._lock:
            self._items.extend(copy.deepcopy(item) for item in list(items or []))

    async def pop_item(self) -> Optional[Dict[str, Any]]:
        async with self._lock:
            if not self._items:
                return None
            return self._items.pop()

    async def clear_session(self) -> None:
        async with self._lock:
            self._items.clear()


def _acquire_cancellation_event(conversation_id: str) -> Event:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.clear()
        return event


def cancel_conversation_processing(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.set()


def _release_cancellation_event(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        _CANCELLATION_EVENTS.pop(conversation_id, None)


def _ensure_not_cancelled(conversation_id: str, event: Event, stage: str) -> None:
    if event.is_set():
        logger.info("Conversation cancelled cid=%s stage=%s", conversation_id, stage)
        raise ConversationCancelled(f"Conversation {conversation_id} cancelled during {stage}")


def _conversation_runtime_dir(conversation_id: str) -> Path:
    safe_id = "".join(ch for ch in conversation_id if ch.isalnum() or ch in {"-", "_"})
    if not safe_id:
        safe_id = "conversation"
    return settings.runtime_dir / "conversations" / safe_id


def conversation_session(conversation_id: str) -> SQLiteSession:
    return SQLiteSession(conversation_id, db_path=SESSION_DB_PATH)


def _artifact_file_path(conversation_id: str, artifact_id: str) -> Path:
    return _conversation_runtime_dir(conversation_id) / "artifacts" / f"{artifact_id}.json"


def _trace_file_path(conversation_id: str) -> Path:
    return _conversation_runtime_dir(conversation_id) / "agent_trace.jsonl"


def _ensure_runtime_dirs(conversation_id: str) -> Path:
    run_dir = _conversation_runtime_dir(conversation_id)
    (run_dir / "artifacts").mkdir(parents=True, exist_ok=True)
    return run_dir


def _clear_runtime_dir(conversation_id: str) -> None:
    run_dir = _conversation_runtime_dir(conversation_id)
    if run_dir.exists():
        shutil.rmtree(run_dir, ignore_errors=True)


def reset_conversation_runtime(conversation_id: str) -> None:
    _clear_runtime_dir(conversation_id)


def _session_items_from_state_messages(messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for message in messages:
        if not isinstance(message, dict):
            continue
        role = str(message.get("role") or "").strip().lower()
        content = str(message.get("content") or "").strip()
        if not content:
            continue
        if role == "user":
            items.append({"type": "message", "role": "user", "content": content})
        elif role == "assistant":
            items.append(
                {
                    "type": "message",
                    "role": "assistant",
                    "phase": "final_answer",
                    "content": content,
                }
            )
    return items


def _agent_session_items_from_chat_history(
    messages: List[Dict[str, Any]],
    max_pairs: int = 5,
    max_workflow_notes: int = 8,
) -> List[Dict[str, Any]]:
    collected: List[Dict[str, Any]] = []
    workflow_notes: List[str] = []
    for message in messages:
        if not isinstance(message, dict):
            continue
        role = str(message.get("role") or "").strip().lower()
        content = str(message.get("content") or "").strip()
        if not content:
            continue
        if role == "progress":
            workflow_notes.append(_truncate(content, 240))
            continue
        if role not in {"user", "assistant"}:
            continue
        if collected and collected[-1].get("role") == role and collected[-1].get("content") == content:
            continue
        collected.append({"role": role, "content": content})

    workflow_item = None
    recent_workflow_notes = [note for note in workflow_notes[-max(0, max_workflow_notes):] if note]
    if recent_workflow_notes:
        workflow_item = {
            "role": "assistant",
            "content": (
                "Recent workflow notes from earlier runs. Use these for high-level continuity only; "
                "do not treat them as source evidence or raw data:\n- "
                + "\n- ".join(recent_workflow_notes)
            ),
        }

    if len(collected) <= 2:
        return ([workflow_item] if workflow_item else []) + collected

    pairs: List[List[tuple[int, Dict[str, Any]]]] = []
    pending_user: tuple[int, Dict[str, Any]] | None = None
    trailing_items: List[tuple[int, Dict[str, Any]]] = []
    for index, item in enumerate(collected):
        if item["role"] == "user":
            if pending_user is not None:
                trailing_items.append(pending_user)
            pending_user = (index, item)
            continue
        if pending_user is not None:
            pairs.append([pending_user, (index, item)])
            pending_user = None
        else:
            trailing_items.append((index, item))

    if pending_user is not None:
        trailing_items.append(pending_user)

    if pairs:
        selected_pairs = pairs[-max(1, max_pairs):]
        flattened = [entry for pair in selected_pairs for _, entry in pair]
        if trailing_items:
            last_pair_index = selected_pairs[-1][-1][0]
            flattened.extend([item for idx, item in trailing_items if idx > last_pair_index][-2:])
        return ([workflow_item] if workflow_item else []) + flattened

    return ([workflow_item] if workflow_item else []) + collected[-max(2, max_pairs * 2):]


async def clear_agent_session(conversation_id: str) -> None:
    await conversation_session(conversation_id).clear_session()


async def sync_agent_session_from_state(conversation_id: str, state: Any) -> None:
    session = conversation_session(conversation_id)
    await session.clear_session()
    items = _session_items_from_state_messages(list(getattr(state, "messages", []) or []))
    if items:
        await session.add_items(items)


def _truncate(text: Any, length: int = 280) -> str:
    clean = str(text or "").replace("\n", " ").strip()
    return clean if len(clean) <= length else clean[: length - 1] + "…"


def _safe_int(value: Any) -> int:
    try:
        numeric = int(value)
    except Exception:
        return 0
    return numeric if numeric > 0 else 0


def _compute_run_cost_breakdown(
    *,
    input_tokens: int,
    output_tokens: int,
    cached_input_tokens: int = 0,
) -> Dict[str, float]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    uncached_input_tokens = max(input_tokens, 0) - cached_tokens
    ai_cost = (
        (uncached_input_tokens / 1_000_000) * GPT_5_4_INPUT_PRICE_PER_MILLION
        + (cached_tokens / 1_000_000) * GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION
        + (max(output_tokens, 0) / 1_000_000) * GPT_5_4_OUTPUT_PRICE_PER_MILLION
    )
    surcharge = ai_cost * AI_COST_SURCHARGE_RATE
    final_cost = ai_cost + surcharge
    return {
        "ai_cost_usd": round(ai_cost, 6),
        "surcharge_usd": round(surcharge, 6),
        "final_cost_usd": round(final_cost, 6),
    }


def _build_run_cost_payload(
    *,
    input_tokens: int,
    output_tokens: int,
    cached_input_tokens: int = 0,
    model: str | None = None,
) -> Dict[str, Any]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    breakdown = _compute_run_cost_breakdown(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        cached_input_tokens=cached_tokens,
    )
    return {
        "model": str(model or settings.openai_model or "gpt-5.4"),
        "input_tokens": max(input_tokens, 0),
        "cached_input_tokens": cached_tokens,
        "output_tokens": max(output_tokens, 0),
        "pricing": {
            "input_per_million_usd": GPT_5_4_INPUT_PRICE_PER_MILLION,
            "cached_input_per_million_usd": GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION,
            "output_per_million_usd": GPT_5_4_OUTPUT_PRICE_PER_MILLION,
            "surcharge_rate": AI_COST_SURCHARGE_RATE,
        },
        **breakdown,
    }


AGENT_SYSTEM_PROMPT_PATH = PROJECT_ROOT / "AGENT_SYSTEM_PROMPT.md"


def _system_instructions() -> str:
    return AGENT_SYSTEM_PROMPT_PATH.read_text(encoding="utf-8").strip()


@lru_cache(maxsize=1)
def _openai_client() -> OpenAI:
    return OpenAI(api_key=settings.openai_api_key, timeout=settings.openai_timeout_seconds)


def _create_code_container(conversation_id: str) -> str:
    container = _openai_client().containers.create(
        name=f"ausdata-{conversation_id[:24]}",
        memory_limit="1g",
    )
    return str(container.id)


def _build_agent_input(
    user_input: str,
    project_memory: Dict[str, Any] | None,
    model_builder_state: Dict[str, Any] | None,
    pending_validated_variable_candidate: Dict[str, Any] | None = None,
) -> str:
    memory_text = ""
    updated_at = None
    if isinstance(project_memory, dict):
        memory_text = str(project_memory.get("memory_text") or "").strip()
        updated_at = project_memory.get("updated_at")
    context_payload: Dict[str, Any] = {}
    if memory_text:
        context_payload["project_compact_memory"] = {
            "memory_text": memory_text,
            "updated_at": updated_at.isoformat() if hasattr(updated_at, "isoformat") else updated_at,
            "use": "Continuity only; not source evidence.",
        }
    if isinstance(model_builder_state, dict):
        context_payload["project_canvas_context"] = {
            "state": _lightweight_model_builder_state_for_context(model_builder_state),
            "use": "Current visible executable notebook model. Read nodes top-to-bottom, then left-to-right. Use tools for full data, recalculation, refresh, or edits.",
        }
    pending_summary = _pending_validated_variable_candidate_summary(pending_validated_variable_candidate)
    if pending_summary:
        context_payload["pending_validated_variable_candidate"] = {
            **pending_summary,
            "use": "A compiled candidate is waiting for approval/save. If the user approves or says save it, call save_validated_variable without rediscovery or retrieval.",
        }
    if not context_payload:
        return user_input
    return json.dumps(
        {
            "user_message": user_input,
            "project_context": context_payload,
        },
        ensure_ascii=False,
    )


def _pending_validated_variable_candidate_summary(candidate: Dict[str, Any] | None) -> Dict[str, Any]:
    if not isinstance(candidate, dict) or not candidate:
        return {}
    validated_data = candidate.get("validated_data") if isinstance(candidate.get("validated_data"), dict) else {}
    summary: Dict[str, Any] = {
        "compiled": True,
        "name": _truncate(candidate.get("name") or "", 120),
        "node_title": _truncate(candidate.get("node_title") or candidate.get("label") or "", 120),
        "source_name": _truncate(candidate.get("source_name") or "", 120),
        "compiled_at": candidate.get("compiled_at") or "",
        "row_count": validated_data.get("row_count"),
        "columns": validated_data.get("columns") if isinstance(validated_data.get("columns"), list) else [],
    }
    return {key: value for key, value in summary.items() if value not in ("", None, [], {})}


def _compact_node_data_summary(entry: Dict[str, Any]) -> Dict[str, Any]:
    records = entry.get("records") if isinstance(entry.get("records"), list) else []
    columns = [str(item or "").strip() for item in (entry.get("columns") if isinstance(entry.get("columns"), list) else [])]
    series = entry.get("series") if isinstance(entry.get("series"), list) else []
    summary: Dict[str, Any] = {
        "available": bool(entry),
        "kind": entry.get("kind") or "",
        "data_kind": entry.get("data_kind") or "",
        "unit": entry.get("unit") or "",
        "rows": len(records),
        "series": len(series),
        "computed_at": entry.get("computed_at") or "",
    }
    if records:
        try:
            x_index = columns.index("period") if "period" in columns else 0
            y_index = columns.index("value") if "value" in columns else min(1, len(records[0]) - 1)
            first = records[0]
            latest = records[-1]
            if isinstance(first, list) and isinstance(latest, list):
                summary["range"] = {
                    "start": first[x_index] if x_index < len(first) else "",
                    "end": latest[x_index] if x_index < len(latest) else "",
                }
                summary["latest"] = {
                    "x": latest[x_index] if x_index < len(latest) else "",
                    "y": latest[y_index] if y_index < len(latest) else None,
                }
        except Exception:
            pass
    if series:
        series_summaries = []
        all_x_values: List[str] = []
        for item in series[:6]:
            if not isinstance(item, dict):
                continue
            points = item.get("points") if isinstance(item.get("points"), list) else []
            all_x_values.extend(
                str(point.get("x") or "").strip()
                for point in points
                if isinstance(point, dict) and str(point.get("x") or "").strip()
            )
            latest_point = points[-1] if points and isinstance(points[-1], dict) else {}
            series_summaries.append(
                {
                    "name": str(item.get("name") or "").strip(),
                    "points": len(points),
                    "latest": {
                        "x": latest_point.get("x") if isinstance(latest_point, dict) else "",
                        "y": latest_point.get("y") if isinstance(latest_point, dict) else None,
                    },
                }
            )
        if series_summaries:
            summary["series_summary"] = series_summaries
        if all_x_values:
            summary["range"] = {"start": all_x_values[0], "end": all_x_values[-1]}
    return summary


def _lightweight_model_builder_state_for_context(model_builder_state: Dict[str, Any]) -> Dict[str, Any]:
    state = copy.deepcopy(model_builder_state)
    variables = state.get("variables") if isinstance(state.get("variables"), list) else []
    variable_by_id = {
        str(variable.get("id") or "").strip(): variable
        for variable in variables
        if isinstance(variable, dict) and str(variable.get("id") or "").strip()
    }
    node_data = state.get("node_data") if isinstance(state.get("node_data"), dict) else {}
    nodes = state.get("nodes") if isinstance(state.get("nodes"), list) else []
    edges = state.get("edges") if isinstance(state.get("edges"), list) else []
    node_titles = {
        str(node.get("id") or "").strip(): str(node.get("node_title") or node.get("id") or "").strip()
        for node in nodes
        if isinstance(node, dict) and str(node.get("id") or "").strip()
    }

    ordered_nodes = sorted(
        [node for node in nodes if isinstance(node, dict)],
        key=lambda node: (float(node.get("positionY") or 0), float(node.get("positionX") or 0), str(node.get("id") or "")),
    )
    context_nodes: List[Dict[str, Any]] = []
    for node in ordered_nodes:
        node_id = str(node.get("id") or "").strip()
        if not node_id:
            continue
        variable_id = str(node.get("variableId") or "").strip()
        variable = variable_by_id.get(variable_id, {})
        entry = node_data.get(node_id) if isinstance(node_data.get(node_id), dict) else {}
        inputs = [
            {"id": input_id, "title": node_titles.get(input_id, input_id)}
            for input_id in ([str(item).strip() for item in node.get("inputs", [])] if isinstance(node.get("inputs"), list) else [])
            if input_id
        ]
        if not inputs:
            inputs = [
                {"id": str(edge.get("sourceNodeId") or "").strip(), "title": node_titles.get(str(edge.get("sourceNodeId") or "").strip(), str(edge.get("sourceNodeId") or "").strip())}
                for edge in edges
                if isinstance(edge, dict) and str(edge.get("targetNodeId") or "").strip() == node_id and str(edge.get("sourceNodeId") or "").strip()
            ]
        context_node: Dict[str, Any] = {
            "node_id": node_id,
            "node_title": str(node.get("node_title") or node_id).strip(),
            "type": str(node.get("nodeType") or "variable").strip(),
            "node_description": str(node.get("node_description") or "").strip(),
            "inputs": inputs,
            "node_data": _compact_node_data_summary(entry),
        }
        if variable:
            context_node["variable"] = {
                "id": variable_id,
                "source": variable.get("sourceName") or "",
                "metric": variable.get("metric") or "",
                "unit": variable.get("unit") or "",
                "geography": variable.get("geography") or "",
                "frequency": variable.get("frequency") or "",
                "coverage": " to ".join(
                    item for item in [str(variable.get("periodStart") or "").strip(), str(variable.get("periodEnd") or "").strip()] if item
                ),
            }
        if str(node.get("nodeType") or "").strip() == "calculation":
            context_node["calculation"] = {
                "expression": node.get("expression") or "",
                "method": node.get("method") or "",
                "has_code": bool((node.get("calculationLogic") or {}).get("code")) if isinstance(node.get("calculationLogic"), dict) else False,
            }
        context_nodes.append(context_node)

    return {
        "ordering": "top_to_bottom_then_left_to_right",
        "nodes": context_nodes,
    }


def _schedule_project_memory_compaction(
    *,
    user_id: str,
    project_id: str,
    project_name: str,
    conversation_id: str,
    messages: List[Dict[str, Any]],
) -> None:
    if not user_id or not project_id:
        return

    def run() -> None:
        try:
            saved = compact_project_memory_after_run(
                user_id=user_id,
                project_id=project_id,
                project_name=project_name,
                conversation_id=conversation_id,
                messages=messages,
                model_name=settings.openai_model,
                openai_api_key=settings.openai_api_key,
            )
            logger.info(
                "Project memory compaction finished cid=%s project_id=%s saved=%s",
                conversation_id,
                project_id,
                saved,
            )
        except Exception as exc:
            logger.warning(
                "Project memory compaction failed cid=%s project_id=%s error=%s",
                conversation_id,
                project_id,
                exc,
            )

    Thread(target=run, name=f"nisaba-memory-{conversation_id[:12]}", daemon=True).start()


@function_tool
def report_progress(
    ctx: RunContextWrapper[AgentRuntimeContext],
    message: str,
) -> Dict[str, Any]:
    """Send one short user-facing progress update.

    Use before/after meaningful steps, saves, graph updates, calculation refreshes, and pivots. Pass one
    factual sentence; do not include hidden reasoning or a long status report.
    """
    normalized = _truncate(message, 220)
    if normalized:
        ctx.context.status_callback(normalized)
    return {"ok": True, "message": normalized}


@function_tool(strict_mode=False)
def search_nisaba_project_memory(
    ctx: RunContextWrapper[AgentRuntimeContext],
    query: str,
    limit: int = 5,
) -> Dict[str, Any]:
    """Search compact memories from the user's other Nisaba projects.

    Use only for continuity about prior project choices, preferences, modelling judgement, or unresolved work.
    Results are not source evidence and must not replace MCP retrieval or saved validated data.
    """
    context = ctx.context
    results = search_project_compact_memory(
        user_id=context.user_id,
        current_project_id=context.project_id,
        query=query,
        limit=limit,
    )
    return {
        "results": results,
        "result_count": len(results),
        "instruction": "Continuity only. Use saved variables, model cache, MCP retrieval, or public-source evidence for facts and data.",
    }


@function_tool(strict_mode=False)
def list_validated_variables(
    ctx: RunContextWrapper[AgentRuntimeContext],
    query: str = "",
    limit: int = 25,
) -> Dict[str, Any]:
    """List validated variables so the agent can identify an existing target.

    Use when the project context does not clearly identify the saved variable to inspect, update, replace,
    duplicate, delete, or disambiguate. Results include active_project_count and active_projects; confirm
    before editing a shared variable. If intent remains unclear, ask one short question.
    """
    context = ctx.context
    return list_validated_variable_records(
        user_id=context.user_id,
        project_id=context.project_id,
        query=query,
        limit=limit,
    )


def _custom_data_official_source_match(
    *,
    source_name: str,
    provider_id: str,
    dataset_id: str,
    validated_api_url: str,
    retrieval_logic: Dict[str, Any],
    custom_data: Dict[str, Any],
) -> str:
    text = "\n".join(
        str(part or "").lower()
        for part in (
            source_name,
            provider_id,
            dataset_id,
            validated_api_url,
            json.dumps(retrieval_logic, ensure_ascii=False, default=str),
            json.dumps(custom_data, ensure_ascii=False, default=str),
        )
    )
    tokens = set(re.split(r"[^a-z0-9]+", text))
    for label, markers in SUPPORTED_OFFICIAL_CUSTOM_DATA_SOURCES.items():
        for marker in markers:
            marker_text = marker.lower()
            if " " in marker_text or "." in marker_text:
                if marker_text in text:
                    return label
            elif marker_text in tokens:
                return label
    return ""


def _assert_custom_data_path_is_allowed(
    *,
    source_name: str,
    provider_id: str,
    dataset_id: str,
    validated_api_url: str,
    retrieval_logic: Dict[str, Any],
    custom_data: Dict[str, Any],
) -> None:
    matched_source = _custom_data_official_source_match(
        source_name=source_name,
        provider_id=provider_id,
        dataset_id=dataset_id,
        validated_api_url=validated_api_url,
        retrieval_logic=retrieval_logic,
        custom_data=custom_data,
    )
    if matched_source:
        raise RuntimeError(
            f"Do not save {matched_source} data as research/custom-only data. If the approved output is derived "
            "from official source slices, pass custom_data as the approved final data plus evidence_artifact "
            "pointing to the already narrowed MCP artifacts and executable transformation_logic.code."
        )


def _compile_validated_variable_candidate_payload(
    *,
    context: AgentRuntimeContext,
    name: str,
    node_title: str = "",
    source_name: str = "",
    provider_id: str = "",
    dataset_id: str = "",
    metric: str = "",
    unit: str = "",
    geography: str = "",
    frequency: str = "",
    seasonal_treatment: str = "",
    period_start: str = "",
    period_end: str = "",
    validated_api_url: str = "",
    retrieval_logic: Optional[Dict[str, Any]] = None,
    transformation_logic: Optional[Dict[str, Any]] = None,
    transform_summary: str = "",
    recreation_summary: str = "",
    node_description: str = "",
    custom_data: Optional[Dict[str, Any]] = None,
    evidence_artifact: Optional[Dict[str, Any]] = None,
    external_key: str = "",
) -> Dict[str, Any]:
    retrieval_logic = retrieval_logic or {}
    transformation_logic = _with_default_identity_transform(transformation_logic or {})
    custom_data = custom_data or {}
    evidence_artifact = evidence_artifact or {}
    official_source = _custom_data_official_source_match(
        source_name=source_name,
        provider_id=provider_id,
        dataset_id=dataset_id,
        validated_api_url=validated_api_url,
        retrieval_logic=retrieval_logic,
        custom_data=custom_data,
    )
    validate_refresh_execution = not official_source
    if custom_data and official_source:
        source_records = _narrowed_artifact_records_from_inputs(
            context,
            retrieval_logic,
            evidence_artifact,
            include_all_if_none=True,
        )
        if not source_records:
            raise RuntimeError(
                "Cannot save official-derived variable: no already narrowed MCP source artifacts are available. "
                "Reuse the artifacts from the chart you just built; do not reretrieve only to save."
            )
        retrieval_logic, evidence_artifact = _executed_validated_recipe_from_narrowed_records(context, source_records)
        validated_api_url, validated_api_urls = _validated_api_urls_from_mcp_artifacts(
            context=context,
            current_url=validated_api_url,
            dataset_id=dataset_id,
            retrieval_logic=retrieval_logic,
            evidence_artifact=evidence_artifact,
        )
        retrieval_logic = _stamp_validated_api_url_on_retrieve_steps(retrieval_logic, validated_api_url, validated_api_urls)
        validated_data = _compact_official_derived_data_from_custom_data(
            variable_name=name or node_title,
            custom_data=custom_data,
            transformation_logic=transformation_logic,
            source={
                "source_name": official_source,
                "api_request_urls": validated_api_urls,
                "artifact_trail": evidence_artifact.get("artifact_trail") if isinstance(evidence_artifact.get("artifact_trail"), list) else [],
            },
        )
        _assert_official_derived_transform_matches_approved_data(
            context=context,
            records=source_records,
            variable_name=name or node_title,
            transformation_logic=transformation_logic,
            approved_data=validated_data,
        )
        refresh_metadata = _refresh_metadata_from_retrieval_logic(
            name=name or node_title,
            validated_api_url=validated_api_url,
            retrieval_logic=retrieval_logic,
            transformation_logic=transformation_logic,
            transform_summary=transform_summary,
            recreation_summary=recreation_summary,
        )
        refresh_code = _refresh_code_from_metadata(refresh_metadata)
    elif custom_data:
        _assert_custom_data_path_is_allowed(
            source_name=source_name,
            provider_id=provider_id,
            dataset_id=dataset_id,
            validated_api_url=validated_api_url,
            retrieval_logic=retrieval_logic,
            custom_data=custom_data,
        )
        retrieval_logic = _research_retrieval_logic_from_custom_data(
            name=name or node_title,
            retrieval_logic=retrieval_logic,
            custom_data=custom_data,
        )
        evidence_artifact = _research_evidence_from_custom_data(custom_data)
        validated_data = _compact_validated_data_from_custom_data(
            variable_name=name or node_title,
            custom_data=custom_data,
            transformation_logic=transformation_logic,
        )
        validated_api_url = _research_validated_url(
            name=name or node_title,
            custom_data=custom_data,
            current_url=validated_api_url,
        )
        refresh_metadata = _research_refresh_metadata(
            name=name or node_title,
            validated_api_url=validated_api_url,
            retrieval_logic=retrieval_logic,
            transformation_logic=transformation_logic,
            transform_summary=transform_summary,
            recreation_summary=recreation_summary,
            custom_data=custom_data,
        )
        refresh_code = _research_refresh_code_from_metadata(refresh_metadata, validated_data)
    else:
        retrieval_logic, evidence_artifact = _merge_executed_validated_recipe(
            context=context,
            retrieval_logic=retrieval_logic,
            evidence_artifact=evidence_artifact,
        )
        _assert_validated_variable_is_narrowed(context, retrieval_logic, evidence_artifact)
        narrowed_record = _narrowed_artifact_record_from_inputs(context, retrieval_logic, evidence_artifact)
        if not isinstance(narrowed_record, dict):
            raise RuntimeError("Cannot save validated variable: no exact narrowed artifact is selected for compact data storage.")
        validated_data = _compact_validated_data_from_artifact_record(
            context=context,
            record=narrowed_record,
            variable_name=name or node_title,
            transformation_logic=transformation_logic,
        )
        validated_api_url, validated_api_urls = _validated_api_urls_from_mcp_artifacts(
            context=context,
            current_url=validated_api_url,
            dataset_id=dataset_id,
            retrieval_logic=retrieval_logic,
            evidence_artifact=evidence_artifact,
        )
        retrieval_logic = _stamp_validated_api_url_on_retrieve_steps(retrieval_logic, validated_api_url, validated_api_urls)
        refresh_metadata = _refresh_metadata_from_retrieval_logic(
            name=name or node_title,
            validated_api_url=validated_api_url,
            retrieval_logic=retrieval_logic,
            transformation_logic=transformation_logic,
            transform_summary=transform_summary,
            recreation_summary=recreation_summary,
        )
        refresh_code = _refresh_code_from_metadata(refresh_metadata)
    return {
        "name": name,
        "label": node_title,
        "node_title": node_title,
        "source_name": source_name,
        "provider_id": provider_id,
        "dataset_id": dataset_id,
        "metric": metric,
        "unit": unit,
        "geography": geography,
        "frequency": frequency,
        "seasonal_treatment": seasonal_treatment,
        "period_start": period_start,
        "period_end": period_end,
        "validated_api_url": validated_api_url,
        "retrieval_logic": retrieval_logic,
        "transformation_logic": transformation_logic,
        "transform_summary": transform_summary,
        "recreation_summary": recreation_summary,
        "node_description": node_description,
        "validated_data": validated_data,
        "refresh_code": refresh_code,
        "refresh_metadata": refresh_metadata,
        "evidence_artifact": evidence_artifact,
        "external_key": external_key,
        "validate_refresh_execution": validate_refresh_execution,
    }


@function_tool(strict_mode=False)
def compile_validated_variable_candidate(
    ctx: RunContextWrapper[AgentRuntimeContext],
    name: str,
    node_title: str = "",
    source_name: str = "",
    provider_id: str = "",
    dataset_id: str = "",
    metric: str = "",
    unit: str = "",
    geography: str = "",
    frequency: str = "",
    seasonal_treatment: str = "",
    period_start: str = "",
    period_end: str = "",
    validated_api_url: str = "",
    retrieval_logic: Optional[Dict[str, Any]] = None,
    transformation_logic: Optional[Dict[str, Any]] = None,
    transform_summary: str = "",
    recreation_summary: str = "",
    node_description: str = "",
    custom_data: Optional[Dict[str, Any]] = None,
    evidence_artifact: Optional[Dict[str, Any]] = None,
    external_key: str = "",
) -> Dict[str, Any]:
    """Compile the current approved candidate variable from MCP artifacts and transform code.

    Use after MCP discover/retrieve/inspect/narrow/transform and before asking the user to approve/save a
    variable. This is the final MCP workflow step: it reuses current narrowed artifacts, stores the approved
    compact data, executable refresh_code, retrieval_logic, transformation_logic, metadata, and node text as
    a pending package. Do not reretrieve source data to compile.
    """
    context = ctx.context
    compiled = _compile_validated_variable_candidate_payload(
        context=context,
        name=name,
        node_title=node_title,
        source_name=source_name,
        provider_id=provider_id,
        dataset_id=dataset_id,
        metric=metric,
        unit=unit,
        geography=geography,
        frequency=frequency,
        seasonal_treatment=seasonal_treatment,
        period_start=period_start,
        period_end=period_end,
        validated_api_url=validated_api_url,
        retrieval_logic=retrieval_logic,
        transformation_logic=transformation_logic,
        transform_summary=transform_summary,
        recreation_summary=recreation_summary,
        node_description=node_description,
        custom_data=custom_data,
        evidence_artifact=evidence_artifact,
        external_key=external_key,
    )
    state = context.store.load(context.conversation_id)
    state.pending_validated_variable_candidate = {
        **_coerce_jsonable(compiled),
        "compiled_at": datetime.now(timezone.utc).isoformat(),
        "source_artifact_ids": _artifact_ids_from_value(compiled.get("evidence_artifact")),
    }
    context.store.save(state)
    return {
        "compiled": True,
        "name": compiled.get("name"),
        "node_title": compiled.get("node_title") or compiled.get("label"),
        "row_count": (compiled.get("validated_data") or {}).get("row_count"),
        "columns": (compiled.get("validated_data") or {}).get("columns"),
        "instruction": "Candidate compiled. If the user approves, call save_validated_variable with no reretrieval.",
    }


@function_tool(strict_mode=False)
def save_validated_variable(
    ctx: RunContextWrapper[AgentRuntimeContext],
    name: str = "",
    node_title: str = "",
    source_name: str = "",
    provider_id: str = "",
    dataset_id: str = "",
    metric: str = "",
    unit: str = "",
    geography: str = "",
    frequency: str = "",
    seasonal_treatment: str = "",
    period_start: str = "",
    period_end: str = "",
    validated_api_url: str = "",
    retrieval_logic: Optional[Dict[str, Any]] = None,
    transformation_logic: Optional[Dict[str, Any]] = None,
    transform_summary: str = "",
    recreation_summary: str = "",
    node_description: str = "",
    custom_data: Optional[Dict[str, Any]] = None,
    evidence_artifact: Optional[Dict[str, Any]] = None,
    external_key: str = "",
    update_variable_id: str = "",
    allow_shared_update: bool = False,
) -> Dict[str, Any]:
    """Persist the approved compiled validated-variable candidate.

    Prefer saving the pending package from compile_validated_variable_candidate. Use full arguments only when
    compiling and saving in one call. Never reretrieve source data just to save after the user approves.
    """
    context = ctx.context
    state = context.store.load(context.conversation_id)
    has_supplied_candidate = bool(custom_data or retrieval_logic or evidence_artifact)
    pending = state.pending_validated_variable_candidate if isinstance(state.pending_validated_variable_candidate, dict) else {}
    if pending and not has_supplied_candidate:
        compiled = dict(pending)
        if name:
            compiled["name"] = name
        if node_title:
            compiled["label"] = node_title
            compiled["node_title"] = node_title
        if node_description:
            compiled["node_description"] = node_description
    elif not has_supplied_candidate:
        raise RuntimeError(
            "No compiled validated-variable candidate is pending. Compile the approved candidate from the current MCP artifacts before saving."
        )
    else:
        compiled = _compile_validated_variable_candidate_payload(
            context=context,
            name=name,
            node_title=node_title,
            source_name=source_name,
            provider_id=provider_id,
            dataset_id=dataset_id,
            metric=metric,
            unit=unit,
            geography=geography,
            frequency=frequency,
            seasonal_treatment=seasonal_treatment,
            period_start=period_start,
            period_end=period_end,
            validated_api_url=validated_api_url,
            retrieval_logic=retrieval_logic,
            transformation_logic=transformation_logic,
            transform_summary=transform_summary,
            recreation_summary=recreation_summary,
            node_description=node_description,
            custom_data=custom_data,
            evidence_artifact=evidence_artifact,
            external_key=external_key,
        )
    result = save_validated_variable_record(
        user_id=context.user_id,
        project_id=context.project_id,
        name=str(compiled.get("name") or ""),
        label=str(compiled.get("label") or compiled.get("node_title") or ""),
        source_name=str(compiled.get("source_name") or ""),
        provider_id=str(compiled.get("provider_id") or ""),
        dataset_id=str(compiled.get("dataset_id") or ""),
        metric=str(compiled.get("metric") or ""),
        unit=str(compiled.get("unit") or ""),
        geography=str(compiled.get("geography") or ""),
        frequency=str(compiled.get("frequency") or ""),
        seasonal_treatment=str(compiled.get("seasonal_treatment") or ""),
        period_start=str(compiled.get("period_start") or ""),
        period_end=str(compiled.get("period_end") or ""),
        validated_api_url=str(compiled.get("validated_api_url") or ""),
        retrieval_logic=compiled.get("retrieval_logic") if isinstance(compiled.get("retrieval_logic"), dict) else {},
        transformation_logic=compiled.get("transformation_logic") if isinstance(compiled.get("transformation_logic"), dict) else {},
        transform_summary=str(compiled.get("transform_summary") or ""),
        recreation_summary=str(compiled.get("recreation_summary") or ""),
        node_description=str(compiled.get("node_description") or ""),
        validated_data=compiled.get("validated_data") if isinstance(compiled.get("validated_data"), dict) else {},
        refresh_code=str(compiled.get("refresh_code") or ""),
        refresh_metadata=compiled.get("refresh_metadata") if isinstance(compiled.get("refresh_metadata"), dict) else {},
        evidence_artifact=compiled.get("evidence_artifact") if isinstance(compiled.get("evidence_artifact"), dict) else {},
        external_key=str(compiled.get("external_key") or external_key or ""),
        update_variable_id=update_variable_id,
        allow_shared_update=allow_shared_update,
        validate_refresh_execution=bool(compiled.get("validate_refresh_execution", True)),
    )
    if result.get("saved"):
        state = context.store.load(context.conversation_id)
        state.pending_validated_variable_candidate = None
        context.store.save(state)
        _clear_heavy_validation_context(context)
    return result


@function_tool(strict_mode=False)
def run_validated_variable(
    ctx: RunContextWrapper[AgentRuntimeContext],
    variable_id: str = "",
    external_key: str = "",
    name: str = "",
    refresh: bool = False,
) -> Dict[str, Any]:
    """Return saved compact data for an active validated variable by id, external_key, or name.

    Use refresh=False to inspect, show, or calculate from stored validated_data only. Do not describe this
    as replaying and do not rewrite state. Use refresh=True only to rerun the same approved refresh_code
    from source; changing scope, history, frequency, filters, transformation, notes, or definition is a
    variable revision that must be re-approved and saved.
    """
    context = ctx.context
    return run_validated_variable_record(
        user_id=context.user_id,
        project_id=context.project_id,
        conversation_id=context.conversation_id,
        code_container_id=context.code_container_id,
        variable_id=variable_id,
        external_key=external_key,
        name=name,
        refresh=refresh,
    )


@function_tool(strict_mode=False)
def run_model_node_calculation(
    ctx: RunContextWrapper[AgentRuntimeContext],
    node_id: str,
    refresh: bool = False,
) -> Dict[str, Any]:
    """Return or refresh saved node_data for a model graph node.

    Use refresh=False to inspect the saved node_data. Use refresh=True only when the user asks to refresh
    a calculation or after changing a node's inputs/calculation code. Missing node_data for a visible node
    is a model integrity error; fix the save/update that failed to write it.
    """
    context = ctx.context
    return build_model_node_chart(
        user_id=context.user_id,
        project_id=context.project_id,
        node_id=node_id,
        refresh=refresh,
    )


@function_tool(strict_mode=False)
def save_custom_calculation(
    ctx: RunContextWrapper[AgentRuntimeContext],
    node_id: str,
    node_title: str,
    input_node_ids: List[str],
    output_label: str = "",
    expression: str = "",
    method: str = "",
    node_description: str = "",
    calculation_logic: Optional[Dict[str, Any]] = None,
    parameters: Optional[Dict[str, Any]] = None,
    position_x: Optional[float] = None,
    position_y: Optional[float] = None,
) -> Dict[str, Any]:
    """Save a custom calculation node, visible description, executable logic, and node_data.

    Use for projections, scenarios, judgement-based transformations, non-obvious calculations, and named
    arithmetic outputs. Pass node_id, node_title, node_description, exact input node ids, parameters, and
    calculation_logic.code. calculation_logic.code must define calculate(inputs, parameters) and return
    chartable points. node_description is the single explanation the user and model both see; it must embed
    the calculation and any modelling judgement together. Aim for 50 words, use up to 75 when needed, never
    more than 100. Do not create standalone symbol nodes; the named calculation node is the output and owns
    its inputs/expression/node_data. Saving runs the calculation immediately and stores chart-ready node_data.
    """
    context = ctx.context
    previous_state = fetch_model_builder_state(user_id=context.user_id, project_id=context.project_id)
    try:
        result = save_custom_calculation_state(
            user_id=context.user_id,
            project_id=context.project_id,
            node_id=node_id,
            node_title=node_title,
            input_node_ids=input_node_ids,
            output_label=output_label,
            expression=expression,
            method=method,
            node_description=node_description,
            calculation_logic=calculation_logic,
            parameters=parameters,
            position_x=position_x,
            position_y=position_y,
        )
        node_data = build_model_node_chart(
            user_id=context.user_id,
            project_id=context.project_id,
            node_id=result["calculation_node_id"],
            refresh=True,
        )
    except Exception:
        if previous_state:
            update_model_graph_state(
                user_id=context.user_id,
                project_id=context.project_id,
                variables=previous_state.get("variables") if isinstance(previous_state.get("variables"), list) else [],
                nodes=previous_state.get("nodes") if isinstance(previous_state.get("nodes"), list) else [],
                edges=previous_state.get("edges") if isinstance(previous_state.get("edges"), list) else [],
            )
        raise
    return {
        "updated": True,
        **result,
        "node_data": node_data,
        "instruction": (
            "The custom calculation is saved as one graph node with replayable calculation details, "
            "a concise node_description, and saved node_data."
        ),
    }


@function_tool(strict_mode=False)
def update_model_builder(
    ctx: RunContextWrapper[AgentRuntimeContext],
    model_builder_state: Dict[str, Any],
) -> Dict[str, Any]:
    """Replace the complete model-builder state: variables, nodes, and edges.

    Use only when a full replacement is intended. Prefer update_model_graph for graph-only edits.
    Preserve existing stable node ids and all valid links unless the user explicitly asks to restructure the
    model. Every visible node must include node_title, node_description, valid edges, and saved node_data.
    """
    context = ctx.context
    result = update_model_builder_state(
        user_id=context.user_id,
        project_id=context.project_id,
        model_builder_state=model_builder_state,
    )
    return {
        "updated": True,
        **result,
        "instruction": "The right-pane model builder state has been updated for this project.",
    }


@function_tool(strict_mode=False)
def update_model_graph(
    ctx: RunContextWrapper[AgentRuntimeContext],
    nodes: List[Dict[str, Any]],
    edges: List[Dict[str, Any]],
    variables: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Replace only the model graph and, when supplied, active validated-variable links.

    Use for graph-only edits: visible nodes, edges, and optional active variable links. Treat links as core
    executable model structure. Preserve existing stable node ids, node_data ids, inputs, and edges unless
    the user explicitly asks to restructure those relationships. Never replace semantic ids with generated
    node-1/node-2 style ids. Calculation/output nodes must include exact upstream input ids, and edges must
    mirror those inputs. Simple arithmetic may use expression plus inputs on the named output node itself.
    Projections, annualisation, filters, transformations, scenarios, and aggregations need
    calculationLogic.code defining calculate(inputs, parameters). Every visible node must include
    node_title and node_description. For every calculation node, node_description must state both what
    the node calculates and the assumption that makes that calculation valid. If the same assumption
    applies across multiple nodes, reuse the same wording for that assumption. Aim for 50 words, use up
    to 75 when needed, never more than 100.
    Do not create standalone symbol nodes. If graph changes alter executable calculation logic, refresh the
    affected calculation node so its node_data is current.
    """
    context = ctx.context
    result = update_model_graph_state(
        user_id=context.user_id,
        project_id=context.project_id,
        variables=variables,
        nodes=nodes,
        edges=edges,
    )
    return {
        "updated": True,
        **result,
        "instruction": "The right-pane model graph has been updated for this project.",
    }


_TRANSIENT_MODEL_STATUSES = {408, 409, 429, 500, 502, 503, 504}
_RETRY_DELAY_BY_ATTEMPT = {
    1: 0.5,
    2: 1.5,
    3: 5.0,
    4: 12.0,
}


def _model_retry_policy(context: RetryPolicyContext) -> bool | RetryDecision:
    advice = context.provider_advice
    if advice is not None and advice.suggested is False:
        return RetryDecision(retry=False, reason=advice.reason)

    normalized = context.normalized
    status_code = normalized.status_code
    should_retry = (
        normalized.is_network_error
        or normalized.is_timeout
        or (status_code in _TRANSIENT_MODEL_STATUSES if status_code is not None else False)
        or bool(advice and advice.suggested)
    )
    if not should_retry:
        return False

    delay = _RETRY_DELAY_BY_ATTEMPT.get(context.attempt)
    provider_delay = advice.retry_after if advice is not None else normalized.retry_after
    if provider_delay is not None:
        delay = max(delay or 0.0, float(provider_delay))

    logger.warning(
        "Model retry scheduled attempt=%s/%s status=%s network=%s timeout=%s delay_s=%s request_id=%s reason=%s",
        context.attempt,
        context.max_retries,
        status_code,
        normalized.is_network_error,
        normalized.is_timeout,
        delay,
        normalized.request_id,
        advice.reason if advice is not None else normalized.message,
    )
    return RetryDecision(retry=True, delay=delay)


def _build_agent(code_container_id: str) -> Agent[Any]:
    return Agent(
        name="AusData",
        model=settings.openai_model,
        instructions=_system_instructions(),
        tools=[
            report_progress,
            search_nisaba_project_memory,
            list_validated_variables,
            compile_validated_variable_candidate,
            save_validated_variable,
            run_validated_variable,
            run_model_node_calculation,
            save_custom_calculation,
            update_model_builder,
            update_model_graph,
            WebSearchTool(search_context_size="medium"),
            CodeInterpreterTool(
                tool_config={
                    "type": "code_interpreter",
                    "container": code_container_id,
                }
            ),
        ],
        model_settings=ModelSettings(
            reasoning={"effort": settings.openai_reasoning_effort},
            include_usage=True,
            parallel_tool_calls=False,
            retry=ModelRetrySettings(
                max_retries=4,
                policy=_model_retry_policy,
                backoff={
                    "initial_delay": 0.5,
                    "max_delay": 12.0,
                    "multiplier": 2.0,
                    "jitter": True,
                },
            ),
        ),
        mcp_config={
            "convert_schemas_to_strict": False,
        },
    )


def _integrated_mcp_server_for_conversation(conversation_id: str, code_container_id: str) -> MCPServerStdio:
    return MCPServerStdio(
        params={
            "command": settings.python_binary,
            "args": ["-m", "backend.app.unified_mcp_server"],
            "cwd": str(PROJECT_ROOT),
            "env": {
                "NISABA_CONVERSATION_ID": conversation_id,
                "NISABA_RUNTIME_DIR": str(settings.runtime_dir),
                "NISABA_CODE_CONTAINER_ID": code_container_id,
                "NODE_BINARY": settings.node_binary,
                "OPENAI_API_KEY": settings.openai_api_key,
            },
        },
        name="ausdata",
        client_session_timeout_seconds=max(120, settings.macro_timeout_seconds),
        cache_tools_list=True,
        tool_filter=create_static_tool_filter(
            allowed_tool_names=["search_catalog", "get_metadata", "retrieve", "inspect_artifact", "narrow_artifact"]
        ),
    )


def _next_artifact_id(artifacts: List[Dict[str, Any]]) -> str:
    return f"artifact-{len(artifacts) + 1:03d}"


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _tool_args_summary(tool_args: Dict[str, Any]) -> Dict[str, Any]:
    summary: Dict[str, Any] = {}
    for key in (
        "query",
        "searchQuery",
        "datasetId",
        "candidateId",
        "dataKey",
        "anchorType",
        "anchorCode",
        "startPeriod",
        "endPeriod",
        "startYear",
        "endYear",
        "detail",
        "dimensionAtObservation",
        "flowCode",
        "frequencyCode",
    ):
        value = tool_args.get(key)
        if value in (None, "", [], {}):
            continue
        if isinstance(value, str):
            summary[key] = _truncate(value, 120)
        else:
            summary[key] = value
    for key in ("dimensionFilters", "dimensionFiltersMap"):
        value = tool_args.get(key)
        if isinstance(value, list) and value:
            summary[key] = value[:6]
            if len(value) > 6:
                summary[f"{key}_count"] = len(value)
        elif isinstance(value, dict) and value:
            clipped: Dict[str, Any] = {}
            for idx, (item_key, item_value) in enumerate(value.items()):
                if idx >= 6:
                    break
                clipped[str(item_key)] = item_value
            summary[key] = clipped
            if len(value) > 6:
                summary[f"{key}_count"] = len(value)
    for key in ("countries", "reporterCodes", "partnerCodes", "hsCodes"):
        value = tool_args.get(key)
        if isinstance(value, list) and value:
            summary[key] = value[:6]
            if len(value) > 6:
                summary[f"{key}_count"] = len(value)
    return summary


def _tool_output_summary(payload: Any) -> Dict[str, Any]:
    if isinstance(payload, dict):
        summary: Dict[str, Any] = {"keys": sorted(str(key) for key in list(payload.keys())[:12])}
        if payload.get("artifact_id"):
            summary["artifact_id"] = _truncate(payload.get("artifact_id"), 120)
        if payload.get("kind"):
            summary["kind"] = _truncate(payload.get("kind"), 120)
        analysis_file = payload.get("analysis_file")
        if isinstance(analysis_file, dict):
            summary["analysis_filename"] = _truncate(analysis_file.get("filename") or "", 120)
        if isinstance(payload.get("dataflows"), list):
            summary["dataflows"] = len(payload["dataflows"])
        if isinstance(payload.get("candidates"), list):
            summary["candidates"] = len(payload["candidates"])
        if isinstance(payload.get("series"), list):
            summary["series"] = len(payload["series"])
        manifest = payload.get("manifest")
        if isinstance(manifest, dict):
            if manifest.get("series_count") is not None:
                summary["series_count"] = manifest.get("series_count")
            if manifest.get("observation_count") is not None:
                summary["observation_count"] = manifest.get("observation_count")
            if manifest.get("point_count") is not None:
                summary["point_count"] = manifest.get("point_count")
        dataset = payload.get("dataset")
        if isinstance(dataset, dict):
            summary["datasetId"] = _truncate(dataset.get("id") or dataset.get("dataset_id") or "", 120)
            summary["datasetName"] = _truncate(dataset.get("name") or "", 120)
        selected = payload.get("selected_indicator")
        if isinstance(selected, dict):
            summary["indicator"] = _truncate(
                selected.get("indicator_label") or selected.get("provider_name") or "",
                120,
            )
        if payload.get("provider"):
            summary["provider"] = _truncate(payload.get("provider"), 120)
        return summary
    if isinstance(payload, list):
        return {"type": "list", "count": len(payload)}
    if isinstance(payload, str):
        return {"type": "text", "preview": _truncate(payload, 160)}
    return {"type": type(payload).__name__}


def _display_tool_args_summary(state, tool_name: str, tool_args: Dict[str, Any]) -> Dict[str, Any]:
    summary = _tool_args_summary(tool_args)
    clean_name = str(tool_name or "").strip().lower()
    if summary:
        return summary
    if clean_name in {"inspect_artifact", "narrow_artifact", "macro_inspect_artifact", "macro_narrow_artifact"}:
        record = _latest_artifact_record(state)
        if record:
            return {
                "artifactId": str(record.get("artifact_id") or "").strip(),
                "kind": str(record.get("kind") or "").strip(),
                "label": _truncate(str(record.get("label") or "").strip(), 120),
                "inferred": "latest_artifact",
            }
    if clean_name == "code_interpreter":
        record = _latest_artifact_record(state)
        if record:
            summary = {
                "artifactId": str(record.get("artifact_id") or "").strip(),
                "kind": str(record.get("kind") or "").strip(),
                "label": _truncate(str(record.get("label") or "").strip(), 120),
            }
            analysis_filename = str(record.get("analysis_filename") or "").strip()
            if analysis_filename:
                summary["analysis_filename"] = _truncate(analysis_filename, 120)
            summary["inferred"] = "latest_analysis_artifact"
            return summary
    return summary


def _event_payload_preview(value: Any, length: int = 400) -> str:
    coerced = _coerce_jsonable(value)
    if coerced is None:
        return ""
    try:
        if isinstance(coerced, (dict, list)):
            return _truncate(json.dumps(coerced, ensure_ascii=False, sort_keys=True), length)
        return _truncate(str(coerced), length)
    except Exception:
        return _truncate(repr(coerced), length)


def _artifact_record_by_id(state, artifact_id: str) -> Optional[Dict[str, Any]]:
    target = str(artifact_id or "").strip()
    if not target:
        return None
    for item in reversed(state.artifacts):
        if not isinstance(item, dict):
            continue
        if str(item.get("artifact_id") or "").strip() == target:
            return item
    return None


def _latest_artifact_record(state) -> Optional[Dict[str, Any]]:
    for item in reversed(state.artifacts):
        if isinstance(item, dict):
            return item
    return None


def _latest_narrowed_artifact_record(state) -> Optional[Dict[str, Any]]:
    for item in reversed(state.artifacts):
        if isinstance(item, dict) and "narrowed" in str(item.get("kind") or "").strip():
            return item
    return None


def _artifact_payload_from_record(record: Dict[str, Any]) -> Any:
    path = Path(str(record.get("path") or ""))
    if not path.exists() or path.suffix.lower() != ".json":
        raise RuntimeError(f"Artifact file is not available for {record.get('artifact_id')}.")
    return json.loads(path.read_text(encoding="utf-8"))


def _is_real_api_url(value: Any) -> bool:
    text = str(value or "").strip()
    if not text.startswith(("https://", "http://")) or "…" in text:
        return False
    return not (text.endswith("...") and not text.endswith("...."))


def _artifact_ids_from_value(value: Any) -> List[str]:
    ids: List[str] = []

    def add(candidate: Any) -> None:
        text = str(candidate or "").strip()
        if text and not text.startswith("${") and text not in ids:
            ids.append(text)

    def walk(item: Any) -> None:
        if isinstance(item, dict):
            for key in ("artifact_id", "artifactId", "parent_artifact_id", "parentArtifactId"):
                add(item.get(key))
            for child in item.values():
                walk(child)
        elif isinstance(item, list):
            for child in item:
                walk(child)

    walk(value)
    return ids


def _artifact_api_request_url(state: Any, record: Dict[str, Any]) -> str:
    direct_url = str(record.get("api_request_url") or "").strip()
    if _is_real_api_url(direct_url):
        return direct_url
    try:
        payload = _artifact_payload_from_record(record)
    except Exception:
        payload = None
    if isinstance(payload, dict):
        payload_url = str(payload.get("api_request_url") or "").strip()
        if _is_real_api_url(payload_url):
            return payload_url
        parent_artifact_id = str(payload.get("parent_artifact_id") or "").strip()
        if parent_artifact_id:
            parent_record = _artifact_record_by_id(state, parent_artifact_id)
            if isinstance(parent_record, dict):
                parent_url = _artifact_api_request_url(state, parent_record)
                if parent_url:
                    return parent_url
    parent_artifact_id = str(record.get("parent_artifact_id") or "").strip()
    if parent_artifact_id:
        parent_record = _artifact_record_by_id(state, parent_artifact_id)
        if isinstance(parent_record, dict):
            return _artifact_api_request_url(state, parent_record)
    return ""


def _artifact_api_request_urls(state: Any, record: Dict[str, Any]) -> List[str]:
    urls: List[str] = []

    def add(value: Any) -> None:
        text = str(value or "").strip()
        if _is_real_api_url(text) and text not in urls:
            urls.append(text)

    for item in record.get("api_request_urls") if isinstance(record.get("api_request_urls"), list) else []:
        add(item)
    add(record.get("api_request_url"))
    try:
        payload = _artifact_payload_from_record(record)
    except Exception:
        payload = None
    if isinstance(payload, dict):
        for item in payload.get("api_request_urls") if isinstance(payload.get("api_request_urls"), list) else []:
            add(item)
        add(payload.get("api_request_url"))
        parent_artifact_id = str(payload.get("parent_artifact_id") or "").strip()
        if parent_artifact_id:
            parent_record = _artifact_record_by_id(state, parent_artifact_id)
            if isinstance(parent_record, dict):
                for item in _artifact_api_request_urls(state, parent_record):
                    add(item)
    parent_artifact_id = str(record.get("parent_artifact_id") or "").strip()
    if parent_artifact_id:
        parent_record = _artifact_record_by_id(state, parent_artifact_id)
        if isinstance(parent_record, dict):
            for item in _artifact_api_request_urls(state, parent_record):
                add(item)
    return urls


def _artifact_dataset_id(record: Dict[str, Any]) -> str:
    try:
        payload = _artifact_payload_from_record(record)
    except Exception:
        return ""
    if not isinstance(payload, dict):
        return ""
    dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
    return str(dataset.get("id") or dataset.get("dataset_id") or payload.get("dataset_id") or "").strip()


def _validated_api_urls_from_mcp_artifacts(
    *,
    context: AgentRuntimeContext,
    current_url: str,
    dataset_id: str,
    retrieval_logic: Dict[str, Any],
    evidence_artifact: Dict[str, Any],
) -> tuple[str, List[str]]:
    state = context.store.load(context.conversation_id)
    for artifact_id in _artifact_ids_from_value(evidence_artifact) + _artifact_ids_from_value(retrieval_logic):
        record = _artifact_record_by_id(state, artifact_id)
        if isinstance(record, dict):
            api_urls = _artifact_api_request_urls(state, record)
            if api_urls:
                return api_urls[0], api_urls

    clean_dataset_id = str(dataset_id or "").strip()
    if clean_dataset_id:
        for record in reversed(list(getattr(state, "artifacts", []) or [])):
            if not isinstance(record, dict):
                continue
            if _artifact_dataset_id(record) != clean_dataset_id:
                continue
            api_urls = _artifact_api_request_urls(state, record)
            if api_urls:
                return api_urls[0], api_urls

    latest_record = _latest_artifact_record(state)
    if isinstance(latest_record, dict):
        api_urls = _artifact_api_request_urls(state, latest_record)
        if api_urls:
            return api_urls[0], api_urls

    if _is_real_api_url(current_url):
        url = str(current_url).strip()
        return url, [url]
    raise RuntimeError("Cannot save validated variable: the MCP retrieve artifact does not include a real api_request_url.")


def _tool_call_from_record_or_trace(context: AgentRuntimeContext, record: Dict[str, Any], expected_tool: str) -> Dict[str, Any]:
    tool_call = record.get("tool_call") if isinstance(record.get("tool_call"), dict) else {}
    if (
        isinstance(tool_call.get("args"), dict)
        and str(tool_call.get("tool_name") or "").strip() == str(expected_tool or "").strip()
    ):
        return tool_call
    artifact_id = str(record.get("artifact_id") or "").strip()
    return _trace_tool_call_for_artifact(context.conversation_id, artifact_id, expected_tool=expected_tool)


def _artifact_evidence_payload(state: Any, record: Dict[str, Any]) -> Dict[str, Any]:
    evidence: Dict[str, Any] = {
        "artifact_id": str(record.get("artifact_id") or "").strip(),
        "parent_artifact_id": str(record.get("parent_artifact_id") or "").strip(),
        "kind": str(record.get("kind") or "").strip(),
        "label": str(record.get("label") or "").strip(),
        "summary": str(record.get("summary") or "").strip(),
    }
    api_url = _artifact_api_request_url(state, record)
    api_urls = _artifact_api_request_urls(state, record)
    if api_url:
        evidence["api_request_url"] = api_url
    if api_urls:
        evidence["api_request_urls"] = api_urls
    refs = record.get("source_references") if isinstance(record.get("source_references"), list) else []
    if refs:
        evidence["source_references"] = refs
    for key in ("analysis_container_id", "analysis_file_id", "analysis_filename", "analysis_local_path"):
        value = str(record.get(key) or "").strip()
        if value:
            evidence[key] = value
    try:
        payload = _artifact_payload_from_record(record)
    except Exception:
        payload = None
    if isinstance(payload, dict):
        evidence["artifact_manifest"] = _artifact_manifest_summary(record, payload)
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
        dataset_id = str(dataset.get("id") or dataset.get("dataset_id") or payload.get("dataset_id") or "").strip()
        if dataset_id:
            evidence["dataset_id"] = dataset_id
    return {key: value for key, value in evidence.items() if value not in ("", [], {})}


def _replace_artifact_arg_with_latest(args: Dict[str, Any]) -> Dict[str, Any]:
    normalized = json.loads(json.dumps(args or {}, ensure_ascii=False))
    if "artifactId" in normalized:
        normalized["artifactId"] = "${latest_artifact_id}"
    elif "artifact_id" in normalized:
        normalized["artifact_id"] = "${latest_artifact_id}"
    else:
        normalized["artifactId"] = "${latest_artifact_id}"
    return normalized


def _has_transformation_details(value: Dict[str, Any]) -> bool:
    return isinstance(value, dict) and any(key in value for key in ("transform_code", "code", "formula", "steps"))


def _with_default_identity_transform(transformation_logic: Dict[str, Any]) -> Dict[str, Any]:
    if _has_transformation_details(transformation_logic):
        return transformation_logic
    normalized = dict(transformation_logic or {})
    normalized["transform_code"] = "return the narrowed series unchanged"
    normalized["transform_type"] = "identity"
    return normalized


def _artifact_chain_evidence_payload(state: Any, narrowed_record: Dict[str, Any], parent_record: Dict[str, Any]) -> Dict[str, Any]:
    narrowed_evidence = _artifact_evidence_payload(state, narrowed_record)
    raw_evidence = _artifact_evidence_payload(state, parent_record)
    artifact_trail = []
    if raw_evidence:
        artifact_trail.append({"role": "retrieve", **raw_evidence})
    if narrowed_evidence:
        artifact_trail.append({"role": "narrow_artifact", **narrowed_evidence})
    if artifact_trail:
        narrowed_evidence["artifact_trail"] = artifact_trail
    return narrowed_evidence


def _executed_validated_recipe_from_narrowed_record(
    context: AgentRuntimeContext,
    narrowed_record: Dict[str, Any],
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    state = context.store.load(context.conversation_id)
    parent_artifact_id = str(narrowed_record.get("parent_artifact_id") or "").strip()
    parent_record = _artifact_record_by_id(state, parent_artifact_id) if parent_artifact_id else None
    if not isinstance(parent_record, dict):
        raise RuntimeError(
            "Cannot save validated variable: the latest narrowed artifact does not link back to its retrieve artifact."
        )

    retrieve_call = _tool_call_from_record_or_trace(context, parent_record, "retrieve")
    narrow_call = _tool_call_from_record_or_trace(context, narrowed_record, "narrow_artifact")
    retrieve_args = retrieve_call.get("args") if isinstance(retrieve_call.get("args"), dict) else {}
    narrow_args = narrow_call.get("args") if isinstance(narrow_call.get("args"), dict) else {}
    if str(retrieve_call.get("tool_name") or "").strip() != "retrieve" or not retrieve_args:
        raise RuntimeError("Cannot save validated variable: the executed retrieve call args are missing from the trace.")
    if str(narrow_call.get("tool_name") or "").strip() != "narrow_artifact" or not narrow_args:
        raise RuntimeError("Cannot save validated variable: the executed narrow_artifact call args are missing from the trace.")

    api_urls = _artifact_api_request_urls(state, narrowed_record)
    inspect_call = _trace_inspect_call_for_artifact(context.conversation_id, parent_artifact_id)
    retrieval_logic: Dict[str, Any] = {
        "version": 1,
        "source": "executed_mcp_recipe",
        "steps": [
            {
                "id": "retrieve",
                "tool": "retrieve",
                "args": json.loads(json.dumps(retrieve_args, ensure_ascii=False)),
            },
            {
                "id": "narrow",
                "tool": "narrow_artifact",
                "args": _replace_artifact_arg_with_latest(narrow_args),
            },
        ],
        "provenance": {
            "raw_artifact_id": str(parent_record.get("artifact_id") or "").strip(),
            "narrowed_artifact_id": str(narrowed_record.get("artifact_id") or "").strip(),
            "retrieve_call_id": str(retrieve_call.get("call_id") or "").strip(),
            "narrow_call_id": str(narrow_call.get("call_id") or "").strip(),
        },
    }
    if api_urls:
        retrieval_logic["validated_api_url"] = api_urls[0]
        retrieval_logic["api_request_urls"] = api_urls
    if inspect_call:
        retrieval_logic["provenance"]["inspect_step"] = {
            "tool": "inspect_artifact",
            "args": inspect_call.get("args") if isinstance(inspect_call.get("args"), dict) else {},
            "call_id": str(inspect_call.get("call_id") or "").strip(),
        }
    for key in ("analysis_container_id", "analysis_file_id", "analysis_filename", "analysis_local_path"):
        value = str(narrowed_record.get(key) or "").strip()
        if value:
            retrieval_logic["provenance"][key] = value

    return retrieval_logic, _artifact_chain_evidence_payload(state, narrowed_record, parent_record)


def _narrowed_artifact_record_from_inputs(
    context: AgentRuntimeContext,
    retrieval_logic: Dict[str, Any],
    evidence_artifact: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    state = context.store.load(context.conversation_id)
    for artifact_id in _artifact_ids_from_value(evidence_artifact) + _artifact_ids_from_value(retrieval_logic):
        record = _artifact_record_by_id(state, artifact_id)
        if isinstance(record, dict) and "narrowed" in str(record.get("kind") or "").strip():
            return record
    return None


def _narrowed_artifact_records_from_inputs(
    context: AgentRuntimeContext,
    retrieval_logic: Dict[str, Any],
    evidence_artifact: Dict[str, Any],
    *,
    include_all_if_none: bool = False,
) -> List[Dict[str, Any]]:
    state = context.store.load(context.conversation_id)
    records: List[Dict[str, Any]] = []

    def add(record: Optional[Dict[str, Any]]) -> None:
        if not isinstance(record, dict) or "narrowed" not in str(record.get("kind") or "").strip():
            return
        artifact_id = str(record.get("artifact_id") or "").strip()
        if artifact_id and all(str(item.get("artifact_id") or "").strip() != artifact_id for item in records):
            records.append(record)

    for artifact_id in _artifact_ids_from_value(evidence_artifact) + _artifact_ids_from_value(retrieval_logic):
        add(_artifact_record_by_id(state, artifact_id))
    if records or not include_all_if_none:
        return records
    for item in getattr(state, "artifacts", []) or []:
        add(item if isinstance(item, dict) else None)
    return records


def _executed_validated_recipe_from_narrowed_records(
    context: AgentRuntimeContext,
    narrowed_records: List[Dict[str, Any]],
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    state = context.store.load(context.conversation_id)
    steps: List[Dict[str, Any]] = []
    artifact_trail: List[Dict[str, Any]] = []
    api_urls: List[str] = []
    provenance_sources: List[Dict[str, str]] = []
    for index, narrowed_record in enumerate(narrowed_records, start=1):
        parent_artifact_id = str(narrowed_record.get("parent_artifact_id") or "").strip()
        parent_record = _artifact_record_by_id(state, parent_artifact_id) if parent_artifact_id else None
        if not isinstance(parent_record, dict):
            raise RuntimeError(
                "Cannot save official-derived variable: every narrowed source artifact must link back to its retrieve artifact."
            )
        retrieve_call = _tool_call_from_record_or_trace(context, parent_record, "retrieve")
        narrow_call = _tool_call_from_record_or_trace(context, narrowed_record, "narrow_artifact")
        retrieve_args = retrieve_call.get("args") if isinstance(retrieve_call.get("args"), dict) else {}
        narrow_args = narrow_call.get("args") if isinstance(narrow_call.get("args"), dict) else {}
        if str(retrieve_call.get("tool_name") or "").strip() != "retrieve" or not retrieve_args:
            raise RuntimeError("Cannot save official-derived variable: an executed retrieve call is missing from the trace.")
        if str(narrow_call.get("tool_name") or "").strip() != "narrow_artifact" or not narrow_args:
            raise RuntimeError("Cannot save official-derived variable: an executed narrow_artifact call is missing from the trace.")
        source_id = f"source_{index}"
        steps.extend(
            [
                {
                    "id": f"retrieve_{index}",
                    "tool": "retrieve",
                    "args": json.loads(json.dumps(retrieve_args, ensure_ascii=False)),
                    "source_id": source_id,
                },
                {
                    "id": f"narrow_{index}",
                    "tool": "narrow_artifact",
                    "args": _replace_artifact_arg_with_latest(narrow_args),
                    "source_id": source_id,
                },
            ]
        )
        for url in _artifact_api_request_urls(state, narrowed_record):
            if url not in api_urls:
                api_urls.append(url)
        artifact_trail.extend(_artifact_chain_evidence_payload(state, narrowed_record, parent_record).get("artifact_trail", []))
        provenance_sources.append(
            {
                "source_id": source_id,
                "raw_artifact_id": str(parent_record.get("artifact_id") or "").strip(),
                "narrowed_artifact_id": str(narrowed_record.get("artifact_id") or "").strip(),
                "retrieve_call_id": str(retrieve_call.get("call_id") or "").strip(),
                "narrow_call_id": str(narrow_call.get("call_id") or "").strip(),
            }
        )
    return (
        {
            "version": 1,
            "source": "executed_mcp_recipe",
            "steps": steps,
            "api_request_urls": api_urls,
            "validated_api_url": api_urls[0] if api_urls else "",
            "provenance": {"sources": provenance_sources},
        },
        {
            "kind": "official_derived",
            "artifact_trail": artifact_trail,
            "api_request_urls": api_urls,
            "api_request_url": api_urls[0] if api_urls else "",
        },
    )


def _executed_validated_recipe_from_latest_narrowed(
    context: AgentRuntimeContext,
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    state = context.store.load(context.conversation_id)
    narrowed_record = _latest_narrowed_artifact_record(state)
    if not isinstance(narrowed_record, dict):
        raise RuntimeError(
            "Cannot save validated variable: no narrowed MCP artifact is available in this conversation. "
            "Run inspect_artifact and narrow_artifact, show the narrowed preview, then save after approval."
        )
    return _executed_validated_recipe_from_narrowed_record(context, narrowed_record)


def _merge_executed_validated_recipe(
    *,
    context: AgentRuntimeContext,
    retrieval_logic: Dict[str, Any],
    evidence_artifact: Dict[str, Any],
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    selected_narrowed_record = _narrowed_artifact_record_from_inputs(context, retrieval_logic, evidence_artifact)
    if isinstance(selected_narrowed_record, dict):
        executed_retrieval_logic, executed_evidence_artifact = _executed_validated_recipe_from_narrowed_record(
            context,
            selected_narrowed_record,
        )
    else:
        executed_retrieval_logic, executed_evidence_artifact = _executed_validated_recipe_from_latest_narrowed(context)
    merged_retrieval_logic = dict(executed_retrieval_logic)
    supplied_recreation = str((retrieval_logic or {}).get("recreation_summary") or "").strip()
    if supplied_recreation:
        merged_retrieval_logic["recreation_summary"] = supplied_recreation
    merged_evidence_artifact = dict(executed_evidence_artifact)
    if isinstance(evidence_artifact, dict):
        for key, value in evidence_artifact.items():
            if value not in ("", [], {}, None) and key not in merged_evidence_artifact:
                merged_evidence_artifact[key] = value
    return merged_retrieval_logic, merged_evidence_artifact


def _data_sort_key(row: Dict[str, Any]) -> str:
    for key in ("observationKey", "x", "TIME_PERIOD", "TIME"):
        value = str(row.get(key) or "").strip()
        if value:
            return value
    return ""


def _rows_as_dicts(headers: List[str], rows: List[List[Any]]) -> List[Dict[str, Any]]:
    return [
        {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        for row in rows
    ]


def _custom_data_rows_and_headers(custom_data: Dict[str, Any]) -> tuple[List[str], List[Dict[str, Any]]]:
    rows = custom_data.get("rows")
    if isinstance(rows, list) and all(isinstance(row, dict) for row in rows):
        headers: List[str] = []
        for row in rows:
            for key in row.keys():
                text_key = str(key)
                if text_key not in headers:
                    headers.append(text_key)
        return headers, [dict(row) for row in rows]

    columns = [str(item).strip() for item in custom_data.get("columns", []) if str(item or "").strip()] if isinstance(custom_data.get("columns"), list) else []
    records = custom_data.get("records")
    if columns and isinstance(records, list):
        row_dicts: List[Dict[str, Any]] = []
        for record in records:
            if isinstance(record, dict):
                row_dicts.append({column: record.get(column) for column in columns})
            elif isinstance(record, list):
                row_dicts.append({columns[index]: record[index] if index < len(record) else None for index in range(len(columns))})
        if row_dicts:
            return columns, row_dicts

    points = custom_data.get("points")
    if isinstance(points, list):
        row_dicts = []
        for point in points:
            if isinstance(point, dict):
                x_value = point.get("x", point.get("period", point.get("scenario", point.get("category"))))
                y_value = point.get("y", point.get("value"))
                row_dicts.append({"period": x_value, "value": y_value})
        if row_dicts:
            return ["period", "value"], row_dicts

    series = custom_data.get("series")
    if isinstance(series, list):
        row_dicts = []
        for item in series:
            if not isinstance(item, dict):
                continue
            series_name = str(item.get("name") or item.get("label") or "").strip()
            raw_points = item.get("points")
            if not series_name or not isinstance(raw_points, list):
                continue
            for point in raw_points:
                if isinstance(point, dict):
                    x_value = point.get("x", point.get("period", point.get("year", point.get("category"))))
                    y_value = point.get("y", point.get("value"))
                elif isinstance(point, list):
                    x_value = point[0] if len(point) > 0 else None
                    y_value = point[1] if len(point) > 1 else None
                else:
                    continue
                row_dicts.append({"period": x_value, "series": series_name, "value": y_value})
        if row_dicts:
            return ["period", "series", "value"], row_dicts

    raise RuntimeError("Research-derived variables require custom_data.rows, custom_data.records with columns, custom_data.points, or custom_data.series.")


def _research_evidence_from_custom_data(custom_data: Dict[str, Any]) -> Dict[str, Any]:
    evidence = custom_data.get("evidence") if isinstance(custom_data.get("evidence"), list) else []
    sources = custom_data.get("sources") if isinstance(custom_data.get("sources"), list) else []
    return {
        "kind": "research_derived",
        "label": str(custom_data.get("label") or custom_data.get("name") or "Research-derived variable").strip(),
        "search_queries": custom_data.get("search_queries") if isinstance(custom_data.get("search_queries"), list) else [],
        "sources": sources,
        "evidence": evidence,
        "caveats": custom_data.get("caveats") if isinstance(custom_data.get("caveats"), list) else [],
        "method": str(custom_data.get("method") or "").strip(),
        "confidence": str(custom_data.get("confidence") or "low").strip(),
    }


def _research_retrieval_logic_from_custom_data(
    *,
    name: str,
    retrieval_logic: Dict[str, Any],
    custom_data: Dict[str, Any],
) -> Dict[str, Any]:
    supplied = dict(retrieval_logic) if isinstance(retrieval_logic, dict) else {}
    return {
        **supplied,
        "kind": "research_derived",
        "name": str(name or "").strip(),
        "search_queries": custom_data.get("search_queries") if isinstance(custom_data.get("search_queries"), list) else supplied.get("search_queries", []),
        "source_urls": custom_data.get("source_urls") if isinstance(custom_data.get("source_urls"), list) else supplied.get("source_urls", []),
        "method": str(custom_data.get("method") or supplied.get("method") or "").strip(),
        "confidence": str(custom_data.get("confidence") or supplied.get("confidence") or "low").strip(),
        "caveats": custom_data.get("caveats") if isinstance(custom_data.get("caveats"), list) else supplied.get("caveats", []),
    }


def _compact_validated_data_from_custom_data(
    *,
    variable_name: str,
    custom_data: Dict[str, Any],
    transformation_logic: Dict[str, Any],
) -> Dict[str, Any]:
    headers, rows = _custom_data_rows_and_headers(custom_data)
    if not headers or not rows:
        raise RuntimeError("Research-derived variables require at least one row with reconstructable columns.")
    artifact_payload = {
        "kind": "research_derived",
        "name": variable_name,
        "method": custom_data.get("method"),
        "confidence": custom_data.get("confidence") or "low",
        "evidence": custom_data.get("evidence") if isinstance(custom_data.get("evidence"), list) else [],
        "sources": custom_data.get("sources") if isinstance(custom_data.get("sources"), list) else [],
        "search_queries": custom_data.get("search_queries") if isinstance(custom_data.get("search_queries"), list) else [],
        "caveats": custom_data.get("caveats") if isinstance(custom_data.get("caveats"), list) else [],
    }
    artifact_id = "research_" + hashlib.sha256(json.dumps(_coerce_jsonable(artifact_payload), sort_keys=True).encode("utf-8")).hexdigest()[:16]
    return compact_validated_data_from_rows(
        artifact_kind="research_derived",
        artifact_id=artifact_id,
        variable_name=variable_name,
        headers=headers,
        rows=rows,
        transformation_logic=transformation_logic,
        source=artifact_payload,
    )


def _compact_official_derived_data_from_custom_data(
    *,
    variable_name: str,
    custom_data: Dict[str, Any],
    transformation_logic: Dict[str, Any],
    source: Dict[str, Any],
) -> Dict[str, Any]:
    headers, rows = _custom_data_rows_and_headers(custom_data)
    if not headers or not rows:
        raise RuntimeError("Official-derived variables require approved final custom_data rows, records, points, or series.")
    artifact_payload = {
        "kind": "official_derived",
        "name": variable_name,
        "method": custom_data.get("method"),
        "dimensions": custom_data.get("dimensions") if isinstance(custom_data.get("dimensions"), dict) else {},
        "source": source,
    }
    artifact_id = "official_derived_" + hashlib.sha256(
        json.dumps(_coerce_jsonable(artifact_payload), sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]
    return compact_validated_data_from_rows(
        artifact_kind="official_derived",
        artifact_id=artifact_id,
        variable_name=variable_name,
        headers=headers,
        rows=rows,
        transformation_logic=transformation_logic,
        source=artifact_payload,
    )


def _source_rows_from_narrowed_records(
    *,
    context: AgentRuntimeContext,
    records: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    source_rows: List[Dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        payload = _artifact_payload_from_record(record)
        if not isinstance(payload, dict):
            continue
        kind = str(record.get("kind") or payload.get("kind") or "").strip()
        if kind.startswith("domestic"):
            headers, rows = _flatten_domestic_payload(payload)
        elif kind.startswith("macro"):
            headers, rows = _flatten_macro_payload(payload)
        else:
            raise RuntimeError(f"Cannot save official-derived variable: unsupported narrowed artifact kind '{kind}'.")
        source_id = f"source_{index}"
        for row in _rows_as_dicts([str(header) for header in headers], rows):
            enriched = dict(row)
            enriched["_source_index"] = index
            enriched["_source_step_id"] = source_id
            enriched["_source_artifact_id"] = str(record.get("artifact_id") or "").strip()
            source_rows.append(enriched)
    if not source_rows:
        raise RuntimeError("Cannot save official-derived variable: narrowed source artifacts had no rows.")
    return source_rows


def _assert_official_derived_transform_matches_approved_data(
    *,
    context: AgentRuntimeContext,
    records: List[Dict[str, Any]],
    variable_name: str,
    transformation_logic: Dict[str, Any],
    approved_data: Dict[str, Any],
) -> None:
    transformed_rows = apply_transformation_rows(
        _source_rows_from_narrowed_records(context=context, records=records),
        transformation_logic,
    )
    transformed_headers = list(transformed_rows[0].keys()) if transformed_rows else []
    transformed_data = compact_validated_data_from_rows(
        artifact_kind="official_derived",
        artifact_id=str(approved_data.get("artifact_id") or "official_derived_current_artifacts"),
        variable_name=variable_name,
        headers=transformed_headers,
        rows=transformed_rows,
        transformation_logic=transformation_logic,
        source=approved_data.get("source") if isinstance(approved_data.get("source"), dict) else {},
    )
    expected = {
        "columns": approved_data.get("columns"),
        "records": approved_data.get("records"),
        "dimensions": approved_data.get("dimensions"),
        "period_key": approved_data.get("period_key"),
        "value_key": approved_data.get("value_key"),
        "row_count": approved_data.get("row_count"),
    }
    actual = {
        "columns": transformed_data.get("columns"),
        "records": transformed_data.get("records"),
        "dimensions": transformed_data.get("dimensions"),
        "period_key": transformed_data.get("period_key"),
        "value_key": transformed_data.get("value_key"),
        "row_count": transformed_data.get("row_count"),
    }
    if actual != expected:
        raise RuntimeError(
            "Cannot save official-derived variable: transformation_logic.code does not reproduce the approved chart data "
            "from the already narrowed source artifacts. Reuse the current artifacts and fix the transform code; do not reretrieve."
        )


def _research_validated_url(*, name: str, custom_data: Dict[str, Any], current_url: str = "") -> str:
    if str(current_url or "").strip():
        return str(current_url).strip()
    source_urls = custom_data.get("source_urls") if isinstance(custom_data.get("source_urls"), list) else []
    for url in source_urls:
        text = str(url or "").strip()
        if text.startswith(("http://", "https://")):
            return text
    for key in ("sources", "evidence"):
        items = custom_data.get(key) if isinstance(custom_data.get(key), list) else []
        for item in items:
            if isinstance(item, dict):
                text = str(item.get("url") or "").strip()
                if text.startswith(("http://", "https://")):
                    return text
    slug = re.sub(r"[^a-z0-9]+", "-", str(name or "research-variable").lower()).strip("-")[:80] or "research-variable"
    return f"research://{slug}"


def _research_refresh_metadata(
    *,
    name: str,
    validated_api_url: str,
    retrieval_logic: Dict[str, Any],
    transformation_logic: Dict[str, Any],
    transform_summary: str,
    recreation_summary: str,
    custom_data: Dict[str, Any],
) -> Dict[str, Any]:
    return {
        "version": 1,
        "kind": "research_derived_refresh_metadata",
        "name": str(name or "").strip(),
        "validated_api_url": str(validated_api_url or "").strip(),
        "retrieval_logic": retrieval_logic,
        "transformation_logic": transformation_logic,
        "transform_summary": str(transform_summary or "").strip(),
        "recreation_summary": str(recreation_summary or "").strip(),
        "research": _research_evidence_from_custom_data(custom_data),
    }


def _research_refresh_code_from_metadata(refresh_metadata: Dict[str, Any], validated_data: Dict[str, Any]) -> str:
    metadata_payload = json.dumps(_coerce_jsonable(refresh_metadata), ensure_ascii=False, indent=2, sort_keys=True)
    data_payload = json.dumps(_coerce_jsonable(validated_data), ensure_ascii=False, indent=2, sort_keys=True)
    return f'''"""Refresh code for a Nisaba research-derived validated variable.

This variable was produced from web research and analyst judgement.
The approved compact data is embedded below; RESEARCH_METADATA stores the search queries, source URLs,
extracted figures, caveats/judgement, confidence, and method for the next AI refresh/revision pass.
"""

import json

RESEARCH_METADATA = json.loads({metadata_payload!r})
APPROVED_VALIDATED_DATA = json.loads({data_payload!r})


def refresh(existing_variable=None):
    return APPROVED_VALIDATED_DATA


if __name__ == "__main__":
    print(json.dumps(refresh(), ensure_ascii=False))
'''


def _compact_validated_data_from_artifact_record(
    *,
    context: AgentRuntimeContext,
    record: Dict[str, Any],
    variable_name: str,
    transformation_logic: Dict[str, Any],
) -> Dict[str, Any]:
    payload = _artifact_payload_from_record(record)
    if not isinstance(payload, dict):
        raise RuntimeError("Cannot save validated variable: the approved narrowed artifact payload is unavailable.")
    kind = str(record.get("kind") or "").strip()
    if not kind:
        kind = str(payload.get("kind") or "").strip()
    if kind.startswith("domestic"):
        headers, rows = _flatten_domestic_payload(payload)
    elif kind.startswith("macro"):
        headers, rows = _flatten_macro_payload(payload)
    else:
        raise RuntimeError(f"Cannot save validated variable: unsupported narrowed artifact kind '{kind}'.")
    if not headers or not rows:
        raise RuntimeError("Cannot save validated variable: the approved narrowed artifact has no tabular data.")
    clean_headers = [str(header) for header in headers]
    row_dicts = _rows_as_dicts(clean_headers, rows)
    transformed_rows = apply_transformation_rows(row_dicts, transformation_logic)
    transformed_headers = list(transformed_rows[0].keys()) if transformed_rows else clean_headers
    return compact_validated_data_from_rows(
        artifact_kind=kind,
        artifact_id=str(record.get("artifact_id") or "").strip(),
        variable_name=str(variable_name or "").strip(),
        headers=transformed_headers,
        rows=transformed_rows,
        transformation_logic=transformation_logic,
        source={
            "api_request_url": str(record.get("api_request_url") or "").strip(),
            "api_request_urls": record.get("api_request_urls") if isinstance(record.get("api_request_urls"), list) else [],
        },
    )


def _normalize_refresh_steps(retrieval_logic: Dict[str, Any]) -> List[Dict[str, Any]]:
    steps = retrieval_logic.get("steps") if isinstance(retrieval_logic.get("steps"), list) else []
    normalized: List[Dict[str, Any]] = []
    for index, step in enumerate(steps, start=1):
        if not isinstance(step, dict):
            continue
        normalized.append(
            {
                "id": str(step.get("id") or f"step_{index}").strip(),
                "tool": str(step.get("tool") or "").strip(),
                "args": step.get("args") if isinstance(step.get("args"), dict) else {},
                "source_id": str(step.get("source_id") or "").strip(),
            }
        )
    return normalized


def _refresh_metadata_from_retrieval_logic(
    *,
    name: str = "",
    validated_api_url: str,
    retrieval_logic: Dict[str, Any],
    transformation_logic: Dict[str, Any],
    transform_summary: str,
    recreation_summary: str,
) -> Dict[str, Any]:
    return {
        "version": 1,
        "kind": "validated_variable_refresh_metadata",
        "name": str(name or "").strip(),
        "validated_api_url": str(validated_api_url or "").strip(),
        "api_request_urls": retrieval_logic.get("api_request_urls") if isinstance(retrieval_logic.get("api_request_urls"), list) else [],
        "steps": _normalize_refresh_steps(retrieval_logic),
        "transformation_logic": transformation_logic,
        "transform_summary": str(transform_summary or "").strip(),
        "recreation_summary": str(recreation_summary or "").strip(),
    }


def _refresh_code_from_metadata(refresh_metadata: Dict[str, Any]) -> str:
    payload = json.dumps(_coerce_jsonable(refresh_metadata), ensure_ascii=False, indent=2, sort_keys=True)
    return f'''"""Refresh code for a Nisaba validated variable.

Run from the project backend environment. It recreates the approved narrowed data slice.
"""

import json
from backend.app import unified_mcp_server as mcp_tools
from backend.app.validated_variables import apply_transformation_rows, compact_validated_data_from_rows

REFRESH_METADATA = json.loads({payload!r})


def _interpolate(value, outputs):
    if isinstance(value, str) and value.startswith("${{") and value.endswith("}}"):
        token = value[2:-1].strip()
        if token == "latest_artifact_id":
            return outputs.get("_latest_artifact_id", "")
        return outputs.get(token, "")
    if isinstance(value, dict):
        return {{key: _interpolate(item, outputs) for key, item in value.items()}}
    if isinstance(value, list):
        return [_interpolate(item, outputs) for item in value]
    return value


def _artifact_id(payload):
    if isinstance(payload, dict):
        for key in ("artifact_id", "artifactId", "id"):
            if payload.get(key):
                return str(payload[key])
    return ""


def refresh(existing_variable=None):
    outputs = {{}}
    latest_artifact_id = ""
    source_rows = []
    source_kinds = []
    source_artifact_ids = []
    for step in REFRESH_METADATA["steps"]:
        tool = step["tool"]
        args = _interpolate(step["args"], outputs)
        if tool == "retrieve":
            result = mcp_tools.retrieve(**args)
        elif tool == "narrow_artifact":
            result = mcp_tools.narrow_artifact(**args)
        else:
            raise RuntimeError(f"Unsupported validated-variable refresh tool: {{tool}}")
        latest_artifact_id = _artifact_id(result) or latest_artifact_id
        if latest_artifact_id:
            outputs["_latest_artifact_id"] = latest_artifact_id
        outputs[step["id"]] = result
        if tool == "narrow_artifact":
            payload = mcp_tools._load_artifact_payload(latest_artifact_id)
            kind = mcp_tools._artifact_kind(latest_artifact_id, payload)
            if kind.startswith("domestic"):
                headers, rows = mcp_tools._flatten_domestic_payload(payload)
            elif kind.startswith("macro"):
                headers, rows = mcp_tools._flatten_macro_payload(payload)
            else:
                raise RuntimeError(f"Unsupported validated-variable artifact kind: {{kind}}")
            source_id = step.get("source_id") or step["id"]
            for row in rows:
                row_dict = {{headers[index]: row[index] for index in range(min(len(headers), len(row)))}}
                row_dict["_source_step_id"] = source_id
                row_dict["_source_artifact_id"] = latest_artifact_id
                source_rows.append(row_dict)
            source_kinds.append(kind)
            source_artifact_ids.append(latest_artifact_id)
    if not source_rows:
        raise RuntimeError("Validated-variable refresh did not produce narrowed source rows.")
    transformation_logic = REFRESH_METADATA.get("transformation_logic", {{}})
    transformed_rows = apply_transformation_rows(source_rows, transformation_logic)
    transformed_headers = list(transformed_rows[0].keys()) if transformed_rows else list(source_rows[0].keys())
    return compact_validated_data_from_rows(
        artifact_kind=source_kinds[-1] if len(set(source_kinds)) == 1 else "official_derived",
        artifact_id=source_artifact_ids[-1] if len(source_artifact_ids) == 1 else "official_derived",
        variable_name=REFRESH_METADATA.get("name", ""),
        headers=transformed_headers,
        rows=transformed_rows,
        transformation_logic=transformation_logic,
        source={{
            "validated_api_url": REFRESH_METADATA.get("validated_api_url", ""),
            "api_request_urls": REFRESH_METADATA.get("api_request_urls", []),
        }},
    )


if __name__ == "__main__":
    print(json.dumps(refresh(), ensure_ascii=False))
'''


def _clear_heavy_validation_context(context: AgentRuntimeContext) -> None:
    state = context.store.load(context.conversation_id)
    state.artifacts = []
    state.loop_history = []
    state.current_abs_dataset_shortlist = []
    state.current_macro_indicator_shortlist = []
    state.active_run_artifact_count = 0
    context.store.save(state)
    artifacts_dir = _conversation_runtime_dir(context.conversation_id) / "artifacts"
    if artifacts_dir.exists():
        shutil.rmtree(artifacts_dir, ignore_errors=True)
        artifacts_dir.mkdir(parents=True, exist_ok=True)


def _retrieval_logic_has_narrow_step(value: Any) -> bool:
    if isinstance(value, dict):
        if str(value.get("tool") or "").strip() == "narrow_artifact":
            return True
        return any(_retrieval_logic_has_narrow_step(item) for item in value.values())
    if isinstance(value, list):
        return any(_retrieval_logic_has_narrow_step(item) for item in value)
    return False


def _evidence_references_narrowed_artifact(
    *,
    context: AgentRuntimeContext,
    evidence_artifact: Dict[str, Any],
    retrieval_logic: Dict[str, Any],
) -> bool:
    state = context.store.load(context.conversation_id)
    for artifact_id in _artifact_ids_from_value(evidence_artifact) + _artifact_ids_from_value(retrieval_logic):
        record = _artifact_record_by_id(state, artifact_id)
        if isinstance(record, dict) and "narrowed" in str(record.get("kind") or "").strip():
            return True
    return False


def _assert_validated_variable_is_narrowed(
    context: AgentRuntimeContext,
    retrieval_logic: Dict[str, Any],
    evidence_artifact: Dict[str, Any],
) -> None:
    if _retrieval_logic_has_narrow_step(retrieval_logic) or _evidence_references_narrowed_artifact(
        context=context,
        evidence_artifact=evidence_artifact,
        retrieval_logic=retrieval_logic,
    ):
        return
    raise RuntimeError(
        "Cannot save validated variable from a raw retrieval-only recipe. "
        "Run inspect_artifact, then narrow_artifact to the exact metric/geography/frequency/treatment/period, "
        "and save the validated variable using that narrowed artifact and narrow_artifact step."
    )


def _stamp_validated_api_url_on_retrieve_steps(retrieval_logic: Dict[str, Any], api_url: str, api_urls: List[str]) -> Dict[str, Any]:
    stamped = json.loads(json.dumps(retrieval_logic or {}, ensure_ascii=False))
    if isinstance(stamped.get("steps"), list):
        for step in stamped["steps"]:
            if isinstance(step, dict) and str(step.get("tool") or "").strip() == "retrieve":
                step["validated_api_url"] = api_url
                step["validated_api_urls"] = api_urls
        return stamped
    if isinstance(stamped.get("api_call"), dict):
        stamped["api_call"]["validated_api_url"] = api_url
        stamped["api_call"]["validated_api_urls"] = api_urls
        return stamped
    if str(stamped.get("tool") or "").strip() == "retrieve":
        stamped["validated_api_url"] = api_url
        stamped["validated_api_urls"] = api_urls
    return stamped


def _domestic_preview_rows(payload: Dict[str, Any], limit: int = 8) -> List[Dict[str, Any]]:
    headers, rows = _flatten_domestic_payload(payload)
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return preview


def _macro_preview_rows(payload: Dict[str, Any], limit: int = 8) -> List[Dict[str, Any]]:
    headers, rows = _flatten_macro_payload(payload)
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return preview


def _artifact_manifest_summary(record: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    kind = str(record.get("kind") or "").strip()
    summary: Dict[str, Any] = {
        "artifact_id": str(record.get("artifact_id") or "").strip(),
        "kind": kind,
        "label": str(record.get("label") or "").strip(),
        "summary": str(record.get("summary") or "").strip(),
    }
    if kind.startswith("domestic"):
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
        series_items = payload.get("series") if isinstance(payload.get("series"), list) else []
        observation_count = 0
        dimension_values: Dict[str, List[str]] = {}
        for series in series_items:
            if not isinstance(series, dict):
                continue
            observations = series.get("observations") if isinstance(series.get("observations"), list) else []
            observation_count += len(observations)
            series_dims = series.get("dimensions") if isinstance(series.get("dimensions"), dict) else {}
            for key, value in series_dims.items():
                label = value.get("label") if isinstance(value, dict) else value
                if label is None:
                    continue
                clean = str(label).strip()
                if not clean:
                    continue
                dimension_values.setdefault(str(key), [])
                if clean not in dimension_values[str(key)] and len(dimension_values[str(key)]) < 6:
                    dimension_values[str(key)].append(clean)
        summary.update(
            {
                "dataset_id": str(dataset.get("id") or "").strip(),
                "series_count": len(series_items),
                "observation_count": observation_count,
                "dimensions": dimension_values,
            }
        )
        if kind.endswith("_narrowed"):
            summary["preview_rows"] = _domestic_preview_rows(payload)
    elif kind.startswith("macro"):
        series_items = payload.get("series") if isinstance(payload.get("series"), list) else []
        point_count = 0
        countries: List[str] = []
        frequencies: List[str] = []
        for series in series_items:
            if not isinstance(series, dict):
                continue
            point_count += len(series.get("points") if isinstance(series.get("points"), list) else [])
            country = str(series.get("country_code") or series.get("country") or "").strip()
            if country and country not in countries and len(countries) < 12:
                countries.append(country)
            frequency = str(series.get("frequency") or "").strip()
            if frequency and frequency not in frequencies:
                frequencies.append(frequency)
        summary.update(
            {
                "provider": str(payload.get("provider") or payload.get("provider_key") or "").strip(),
                "series_count": len(series_items),
                "point_count": point_count,
                "countries": countries,
                "frequencies": frequencies,
            }
        )
        if kind.endswith("_narrowed"):
            summary["preview_rows"] = _macro_preview_rows(payload)
    parent_artifact_id = str(record.get("parent_artifact_id") or "").strip()
    if parent_artifact_id:
        summary["parent_artifact_id"] = parent_artifact_id
    analysis_filename = str(record.get("analysis_filename") or "").strip()
    analysis_container_id = str(record.get("analysis_container_id") or "").strip()
    if analysis_filename and analysis_container_id:
        summary["analysis_file"] = {
            "filename": analysis_filename,
            "container_id": analysis_container_id,
            "artifact_id": str(record.get("artifact_id") or "").strip(),
        }
    return summary


def _make_artifact_record(
    *,
    state,
    path: Path,
    kind: str,
    label: str,
    summary: str,
    artifact_id: str | None = None,
    source_references: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    record = {
        "artifact_id": str(artifact_id or "").strip() or _next_artifact_id(state.artifacts),
        "path": str(path),
        "kind": kind,
        "label": label,
        "summary": summary,
    }
    if source_references:
        record["source_references"] = source_references
    state.artifacts.append(record)
    return record


def _coerce_jsonable(value: Any) -> Any:
    if hasattr(value, "model_dump") and callable(value.model_dump):
        try:
            return value.model_dump(mode="json", exclude_none=True)
        except TypeError:
            return value.model_dump()
    if isinstance(value, list):
        return [_coerce_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _coerce_jsonable(item) for key, item in value.items()}
    return value


def _extract_text_output(value: Any) -> str:
    value = _coerce_jsonable(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        fragments: List[str] = []
        for item in value:
            text = _extract_text_output(item)
            if text:
                fragments.append(text)
        return "\n".join(fragment for fragment in fragments if fragment).strip()
    if isinstance(value, dict):
        if value.get("type") == "text" and value.get("text") is not None:
            return str(value.get("text") or "").strip()
        if isinstance(value.get("content"), list):
            return _extract_text_output(value.get("content"))
        if value.get("text") is not None and len(value.keys()) <= 3:
            return str(value.get("text") or "").strip()
    return ""


def _extract_json_payload(value: Any) -> Any:
    value = _coerce_jsonable(value)
    if isinstance(value, dict):
        text = _extract_text_output(value).strip()
        if text:
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
                if match:
                    try:
                        return json.loads(match.group(1))
                    except json.JSONDecodeError:
                        pass
        output_value = value.get("output")
        if output_value is not None:
            nested = _extract_json_payload(output_value)
            if nested is not None:
                return nested
        content_value = value.get("content")
        if content_value is not None:
            nested = _extract_json_payload(content_value)
            if nested is not None:
                return nested
        return value
    if isinstance(value, list):
        if len(value) == 1:
            nested = _extract_json_payload(value[0])
            if nested is not None:
                return nested
        return value
    text = _extract_text_output(value).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
        if match:
            try:
                return json.loads(match.group(1))
            except json.JSONDecodeError:
                return None
    return None


def _extract_raw_item(value: Any) -> Any:
    raw_item = getattr(value, "raw_item", None)
    return raw_item if raw_item is not None else value


def _extract_raw_item_type(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        candidate = raw.get("type")
        return str(candidate).strip() if candidate is not None else ""
    candidate = getattr(raw, "type", None)
    return str(candidate).strip() if candidate is not None else ""


def _extract_mcp_server_label(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        candidate = raw.get("server_label") or raw.get("serverLabel")
        return str(candidate).strip() if candidate is not None else ""
    for key in ("server_label", "serverLabel"):
        candidate = getattr(raw, key, None)
        if candidate is not None:
            return str(candidate).strip()
    return ""


def _extract_call_id(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        for key in ("call_id", "id"):
            item = raw.get(key)
            if isinstance(item, str) and item.strip():
                return item.strip()
        return ""
    for key in ("call_id", "id"):
        item = getattr(raw, key, None)
        if isinstance(item, str) and item.strip():
            return item.strip()
    return ""


def _extract_tool_name(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        for key in ("name", "tool_name"):
            item = raw.get(key)
            if isinstance(item, str) and item.strip():
                return item.strip()
        raw_type = str(raw.get("type") or "").strip().lower()
        if raw_type in {"web_search_call", "web_search_preview"}:
            return "web_search"
        if raw_type in {"code_interpreter_call", "code_interpreter_tool_call"}:
            return "code_interpreter"
        if raw_type in {"mcp_call"}:
            item = raw.get("name")
            if isinstance(item, str) and item.strip():
                return item.strip()
        return ""
    for key in ("name", "tool_name"):
        item = getattr(raw, key, None)
        if isinstance(item, str) and item.strip():
            return item.strip()
    raw_type = str(getattr(raw, "type", "") or "").strip().lower()
    if raw_type in {"web_search_call", "web_search_preview"}:
        return "web_search"
    if raw_type in {"code_interpreter_call", "code_interpreter_tool_call"}:
        return "code_interpreter"
    return ""


def _extract_tool_arguments(value: Any) -> Dict[str, Any]:
    raw = _extract_raw_item(value)
    arguments = None
    if isinstance(raw, dict):
        arguments = raw.get("arguments")
    else:
        arguments = getattr(raw, "arguments", None)
    if isinstance(arguments, dict):
        return arguments
    if isinstance(arguments, str) and arguments.strip():
        try:
            parsed = json.loads(arguments)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    input_payload = getattr(raw, "input", None)
    if isinstance(input_payload, dict):
        return input_payload
    return {}


def _extract_tool_output_payload(value: Any) -> Any:
    direct_output = getattr(value, "output", None)
    payload = _extract_json_payload(direct_output)
    if payload is not None:
        return payload
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        payload = _extract_json_payload(raw.get("output"))
        if payload is not None:
            return payload
    else:
        payload = _extract_json_payload(getattr(raw, "output", None))
        if payload is not None:
            return payload
    return None


def _tool_transport(tool_name: str, item: Any) -> str:
    clean_name = str(tool_name or "").strip().lower()
    raw_type = _extract_raw_item_type(item).lower()
    if clean_name == "web_search" or "web_search" in raw_type:
        return "web"
    if clean_name == "code_interpreter" or "code_interpreter" in raw_type:
        return "code"
    if clean_name.startswith("macro_") or clean_name in {
        "search_catalog",
        "get_metadata",
        "retrieve",
        "inspect_artifact",
        "narrow_artifact",
    }:
        return "mcp"
    if raw_type.startswith("mcp"):
        return "mcp"
    return "tool"


def _append_trace_event(conversation_id: str, payload: Dict[str, Any]) -> None:
    path = _trace_file_path(conversation_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    entry = {"ts": datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z"), **payload}
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False, sort_keys=True) + "\n")


def _read_trace_events(conversation_id: str) -> List[Dict[str, Any]]:
    path = _trace_file_path(conversation_id)
    if not path.exists():
        return []
    events: List[Dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text:
            continue
        try:
            event = json.loads(text)
        except json.JSONDecodeError:
            continue
        if isinstance(event, dict):
            events.append(event)
    return events


def _trace_tool_call_by_call_id(conversation_id: str) -> Dict[str, Dict[str, Any]]:
    calls: Dict[str, Dict[str, Any]] = {}
    for event in _read_trace_events(conversation_id):
        if str(event.get("event") or "").strip() != "tool_called":
            continue
        call_id = str(event.get("call_id") or "").strip()
        if not call_id:
            continue
        calls[call_id] = {
            "call_id": call_id,
            "tool_name": str(event.get("tool_name") or "").strip(),
            "args": event.get("args") if isinstance(event.get("args"), dict) else {},
            "transport": str(event.get("transport") or "").strip(),
            "server_label": str(event.get("server_label") or "").strip(),
        }
    return calls


def _trace_tool_call_for_artifact(conversation_id: str, artifact_id: str, expected_tool: str = "") -> Dict[str, Any]:
    target = str(artifact_id or "").strip()
    if not target:
        return {}
    clean_expected_tool = str(expected_tool or "").strip()
    calls = _trace_tool_call_by_call_id(conversation_id)
    for event in reversed(_read_trace_events(conversation_id)):
        if str(event.get("event") or "").strip() != "artifact_registered":
            continue
        if str(event.get("artifact_id") or "").strip() != target:
            continue
        event_tool_name = str(event.get("tool_name") or "").strip()
        if clean_expected_tool and event_tool_name != clean_expected_tool:
            continue
        call_id = str(event.get("call_id") or "").strip()
        call = calls.get(call_id, {})
        if call:
            return call
        return {
            "call_id": call_id,
            "tool_name": event_tool_name,
            "args": {},
        }
    return {}


def _trace_inspect_call_for_artifact(conversation_id: str, artifact_id: str) -> Dict[str, Any]:
    target = str(artifact_id or "").strip()
    if not target:
        return {}
    for event in reversed(_read_trace_events(conversation_id)):
        if str(event.get("event") or "").strip() != "tool_called":
            continue
        if str(event.get("tool_name") or "").strip() != "inspect_artifact":
            continue
        args = event.get("args") if isinstance(event.get("args"), dict) else {}
        if str(args.get("artifactId") or args.get("artifact_id") or "").strip() == target:
            return {
                "call_id": str(event.get("call_id") or "").strip(),
                "tool_name": "inspect_artifact",
                "args": args,
                "transport": str(event.get("transport") or "").strip(),
                "server_label": str(event.get("server_label") or "").strip(),
            }
    return {}


def _looks_like_domestic_dataset(payload: Any) -> bool:
    return (
        isinstance(payload, dict)
        and isinstance(payload.get("dataset"), dict)
        and isinstance(payload.get("series"), list)
    )


def _looks_like_macro_result(payload: Any) -> bool:
    return (
        isinstance(payload, dict)
        and isinstance(payload.get("series"), list)
        and (
            payload.get("provider")
            or payload.get("provider_key")
            or isinstance(payload.get("selected_indicator"), dict)
        )
        and not isinstance(payload.get("dataset"), dict)
    )


def _persist_retrieval_artifact(
    *,
    state,
    conversation_id: str,
    payload: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    run_dir = _ensure_runtime_dirs(conversation_id)

    if (
        isinstance(payload, dict)
        and str(payload.get("artifact_id") or "").strip()
        and str(payload.get("kind") or "").strip() in {
            "domestic_retrieve",
            "macro_retrieve",
            "domestic_narrowed",
            "macro_narrowed",
        }
    ):
        artifact_id = str(payload.get("artifact_id") or "").strip()
        existing = _artifact_record_by_id(state, artifact_id)
        if existing is not None:
            return existing
        path = _artifact_file_path(conversation_id, artifact_id)
        record = _make_artifact_record(
            state=state,
            path=path,
            kind=str(payload.get("kind") or "").strip(),
            label=str(payload.get("label") or artifact_id).strip(),
            summary=_truncate(str(payload.get("summary") or "").strip() or f"Stored artifact {artifact_id}.", 300),
            artifact_id=artifact_id,
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )
        parent_artifact_id = str(payload.get("parent_artifact_id") or "").strip()
        if parent_artifact_id:
            record["parent_artifact_id"] = parent_artifact_id
        api_request_url = str(payload.get("api_request_url") or "").strip()
        if not api_request_url and parent_artifact_id:
            parent_record = _artifact_record_by_id(state, parent_artifact_id)
            if isinstance(parent_record, dict):
                api_request_url = str(parent_record.get("api_request_url") or "").strip()
        if api_request_url:
            record["api_request_url"] = api_request_url
        api_request_urls = [
            str(item or "").strip()
            for item in (payload.get("api_request_urls") if isinstance(payload.get("api_request_urls"), list) else [])
            if str(item or "").strip()
        ]
        if not api_request_urls and parent_artifact_id:
            parent_record = _artifact_record_by_id(state, parent_artifact_id)
            if isinstance(parent_record, dict) and isinstance(parent_record.get("api_request_urls"), list):
                api_request_urls = [str(item or "").strip() for item in parent_record["api_request_urls"] if str(item or "").strip()]
        if api_request_urls:
            record["api_request_urls"] = api_request_urls
        for key in ("analysis_container_id", "analysis_file_id", "analysis_filename", "analysis_local_path"):
            value = str(payload.get(key) or "").strip()
            if value:
                record[key] = value
        return record

    if _looks_like_domestic_dataset(payload):
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
        label = str(dataset.get("name") or dataset.get("id") or "Domestic dataset").strip()
        artifact_path = run_dir / "artifacts" / f"domestic_retrieve_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="domestic_retrieve",
            label=label,
            summary=_truncate(f"Retrieved domestic dataset '{label}'.", 300),
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )
        api_request_url = str(payload.get("api_request_url") or "").strip()
        if api_request_url:
            record["api_request_url"] = api_request_url
        api_request_urls = [
            str(item or "").strip()
            for item in (payload.get("api_request_urls") if isinstance(payload.get("api_request_urls"), list) else [])
            if str(item or "").strip()
        ]
        if api_request_urls:
            record["api_request_urls"] = api_request_urls
        elif api_request_url:
            record["api_request_urls"] = [api_request_url]
        return record

    if _looks_like_macro_result(payload):
        selected = payload.get("selected_indicator") if isinstance(payload.get("selected_indicator"), dict) else {}
        label = str(
            selected.get("indicator_label")
            or payload.get("concept_label")
            or payload.get("provider")
            or "Macro dataset"
        ).strip()
        artifact_path = run_dir / "artifacts" / f"macro_retrieve_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="macro_retrieve",
            label=label,
            summary=_truncate(f"Retrieved macro dataset '{label}'.", 300),
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )
        api_request_url = str(payload.get("api_request_url") or "").strip()
        if api_request_url:
            record["api_request_url"] = api_request_url
        api_request_urls = [
            str(item or "").strip()
            for item in (payload.get("api_request_urls") if isinstance(payload.get("api_request_urls"), list) else [])
            if str(item or "").strip()
        ]
        if api_request_urls:
            record["api_request_urls"] = api_request_urls
        elif api_request_url:
            record["api_request_urls"] = [api_request_url]
        return record

    return None


def _safe_sheet_text(value: Any) -> str:
    text = str(value or "").strip()
    return text if len(text) <= 32000 else text[:31997] + "..."


def _parse_chart_spec_from_markdown(markdown: str) -> Dict[str, Any] | None:
    text = str(markdown or "").strip()
    if not text:
        return None
    match = re.search(r"```chart\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(1))
    except Exception:
        return None
    if not isinstance(parsed, dict):
        return None
    series = parsed.get("series")
    if not isinstance(series, list) or not series:
        return None
    normalized_series: List[Dict[str, Any]] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        points = entry.get("points")
        if not isinstance(points, list) or not points:
            continue
        normalized_points = []
        for point in points:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            y = point.get("y")
            if not x:
                continue
            try:
                numeric_y = float(y)
            except Exception:
                continue
            normalized_points.append({"x": x, "y": numeric_y})
        if normalized_points:
            normalized_series.append(
                {
                    "name": str(entry.get("name") or "Series").strip() or "Series",
                    "points": normalized_points,
                }
            )
    if not normalized_series:
        return None
    return {
        "type": str(parsed.get("type") or "line").strip() or "line",
        "title": str(parsed.get("title") or "").strip(),
        "xLabel": str(parsed.get("xLabel") or "").strip(),
        "yLabel": str(parsed.get("yLabel") or "").strip(),
        "series": normalized_series,
    }


def _chart_table(chart_spec: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series = chart_spec.get("series") if isinstance(chart_spec.get("series"), list) else []
    if not series:
        return [], []
    x_values: List[str] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        for point in entry.get("points") or []:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            if x and x not in x_values:
                x_values.append(x)
    headers = [str(chart_spec.get("xLabel") or "Metric").strip() or "Metric"]
    headers.extend(str(entry.get("name") or "Series").strip() or "Series" for entry in series if isinstance(entry, dict))
    rows: List[List[Any]] = []
    for x in x_values:
        row: List[Any] = [x]
        for entry in series:
            point_map = {
                str(point.get("x") or "").strip(): point.get("y")
                for point in (entry.get("points") or [])
                if isinstance(point, dict)
            }
            row.append(point_map.get(x))
        rows.append(row)
    return headers, rows


def _safe_sheet_name(value: str, used: set[str]) -> str:
    cleaned = re.sub(r'[:\\/*?\[\]]+', " ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)[:31].strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f" {suffix}"
        candidate = (cleaned[: 31 - len(tail)] + tail).strip()
        suffix += 1
    used.add(candidate)
    return candidate


def _safe_export_filename(user_message: str) -> str:
    words = re.findall(r"[A-Za-z0-9]+", str(user_message or "").lower())
    meaningful = [word for word in words if len(word) > 2][:6]
    stem = "-".join(meaningful) or "analysis-export"
    stem = stem[:64].strip("-") or "analysis-export"
    return f"{stem}.xlsx"


def _write_table(sheet, headers: List[Any], rows: List[List[Any]]) -> None:
    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _flatten_domestic_payload(payload: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series_items = payload.get("series")
    if not isinstance(series_items, list):
        return [], []

    dimension_keys: List[str] = []
    attribute_keys: List[str] = []
    for series in series_items:
        if not isinstance(series, dict):
            continue
        for key in (series.get("dimensions") or {}).keys():
            if key not in dimension_keys:
                dimension_keys.append(str(key))
        for observation in series.get("observations") or []:
            if not isinstance(observation, dict):
                continue
            for key in (observation.get("dimensions") or {}).keys():
                if key not in dimension_keys:
                    dimension_keys.append(str(key))
            for key in (observation.get("attributes") or {}).keys():
                if key not in attribute_keys:
                    attribute_keys.append(str(key))
        for key in (series.get("attributes") or {}).keys():
            if key not in attribute_keys:
                attribute_keys.append(str(key))

    headers = ["seriesKey"] + dimension_keys + ["observationKey", "value"] + attribute_keys
    rows: List[List[Any]] = []

    def _label_or_value(value: Any) -> Any:
        if isinstance(value, dict):
            if value.get("label") is not None:
                return value.get("label")
            if value.get("code") is not None:
                return value.get("code")
        return value

    for series in series_items:
        if not isinstance(series, dict):
            continue
        series_dims = series.get("dimensions") if isinstance(series.get("dimensions"), dict) else {}
        series_attrs = series.get("attributes") if isinstance(series.get("attributes"), dict) else {}
        observations = series.get("observations") if isinstance(series.get("observations"), list) else []
        for observation in observations:
            if not isinstance(observation, dict):
                continue
            obs_dims = observation.get("dimensions") if isinstance(observation.get("dimensions"), dict) else {}
            obs_attrs = observation.get("attributes") if isinstance(observation.get("attributes"), dict) else {}
            row: List[Any] = [series.get("seriesKey")]
            for key in dimension_keys:
                value = obs_dims.get(key)
                if value is None:
                    value = series_dims.get(key)
                row.append(_label_or_value(value))
            row.append(observation.get("observationKey"))
            row.append(observation.get("value"))
            for key in attribute_keys:
                value = obs_attrs.get(key)
                if value is None:
                    value = series_attrs.get(key)
                row.append(_label_or_value(value))
            rows.append(row)

    return headers, rows


def _flatten_macro_payload(payload: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series_items = payload.get("series")
    if not isinstance(series_items, list):
        return [], []
    headers = ["provider", "country", "country_code", "indicator", "series_id", "frequency", "unit", "x", "y"]
    rows: List[List[Any]] = []
    for series in series_items:
        if not isinstance(series, dict):
            continue
        for point in series.get("points") or []:
            if not isinstance(point, dict):
                continue
            rows.append(
                [
                    series.get("provider"),
                    series.get("country"),
                    series.get("country_code"),
                    series.get("indicator"),
                    series.get("series_id"),
                    series.get("frequency"),
                    series.get("unit"),
                    point.get("x"),
                    point.get("y"),
                ]
            )
    return headers, rows


def _write_raw_artifact_sheet(sheet, record: Dict[str, Any]) -> None:
    path = Path(str(record.get("path") or ""))
    payload: Any = None
    if path.exists() and path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
    source_refs = record.get("source_references") if isinstance(record.get("source_references"), list) else []
    source_line = ""
    if source_refs:
        first = source_refs[0]
        if isinstance(first, dict):
            source_line = " | ".join(
                part
                for part in [
                    str(first.get("provider") or "").strip(),
                    str(first.get("dataset_id") or first.get("series_id") or "").strip(),
                    str(first.get("title") or first.get("indicator") or "").strip(),
                    str(first.get("url") or "").strip(),
                ]
                if part
            )
    sheet["A1"] = f"Source: {source_line}".strip() if source_line else "Source data"
    sheet["A2"] = f"Artifact: {str(record.get('label') or record.get('artifact_id') or '').strip()}".strip()
    sheet["A3"] = ""
    sheet["A4"] = "Returned data"

    headers: List[str] = []
    rows: List[List[Any]] = []
    kind = str(record.get("kind") or "").strip()
    if isinstance(payload, dict):
        if kind.startswith("domestic"):
            headers, rows = _flatten_domestic_payload(payload)
        elif kind.startswith("macro"):
            headers, rows = _flatten_macro_payload(payload)

    if headers and rows:
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=col_index, value=header)
        for row_offset, row in enumerate(rows, start=6):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_offset, column=col_index, value=value)
    else:
        raw = json.dumps(payload, ensure_ascii=False, indent=2) if isinstance(payload, (dict, list)) else str(payload or "")
        sheet.cell(row=5, column=1, value=_safe_sheet_text(raw))
        sheet.column_dimensions["A"].width = 140

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")


def _apply_export_theme(workbook) -> None:
    summary = workbook["Summary"] if "Summary" in workbook.sheetnames else None
    if summary is None:
        return
    title_font = Font(color="234233", size=16)
    section_font = Font(color="8F6A3A", size=12)
    header_font = Font(color="54745F", size=11)
    for cell in summary[1]:
        cell.font = title_font
    for row in range(1, summary.max_row + 1):
        first_value = str(summary.cell(row=row, column=1).value or "").strip()
        if first_value in {"Presented data", "Retrieved data"}:
            summary.cell(row=row, column=1).font = section_font
            continue
        values = [str(summary.cell(row=row, column=col).value or "").strip() for col in range(1, summary.max_column + 1)]
        non_empty = [value for value in values if value]
        if len(non_empty) >= 2 and row > 1:
            for col in range(1, len(values) + 1):
                summary.cell(row=row, column=col).font = header_font
            break


def get_latest_export_artifact_path(state) -> Path | None:
    target_id = str(getattr(state, "latest_export_artifact_id", "") or "").strip()
    if not target_id:
        return None
    for item in reversed(state.artifacts):
        if not isinstance(item, dict):
            continue
        if str(item.get("artifact_id") or "").strip() != target_id:
            continue
        path_value = str(item.get("path") or "").strip()
        if not path_value:
            return None
        path = Path(path_value)
        return path if path.exists() else None
    return None


def generate_latest_export(conversation_id: str, store: ConversationStore) -> Path | None:
    state = store.load(conversation_id)
    request = state.latest_export_request if isinstance(state.latest_export_request, dict) else None
    if not request:
        return get_latest_export_artifact_path(state)

    try:
        _build_answer_export(
            state=state,
            conversation_id=conversation_id,
            user_message=str(request.get("user_message") or "").strip(),
            final_answer=str(request.get("final_answer") or "").strip(),
            run_artifact_start_index=int(request.get("run_artifact_start_index") or 0),
        )
        state.latest_export_status = "ready"
        state.latest_export_request = None
        store.save(state)
    except Exception:
        state.latest_export_status = "failed"
        state.latest_export_request = None
        store.save(state)
        raise

    return get_latest_export_artifact_path(state)


def _build_answer_export(
    *,
    state,
    conversation_id: str,
    user_message: str,
    final_answer: str,
    run_artifact_start_index: int,
) -> str:
    chart_spec = _parse_chart_spec_from_markdown(final_answer)
    candidate_artifacts = [
        item
        for item in state.artifacts[run_artifact_start_index:]
        if isinstance(item, dict)
        and str(item.get("kind") or "").strip()
        in {"domestic_retrieve", "macro_retrieve", "domestic_narrowed", "macro_narrowed"}
    ]
    artifact_lookup = {
        str(item.get("artifact_id") or "").strip(): item
        for item in state.artifacts
        if isinstance(item, dict) and str(item.get("artifact_id") or "").strip()
    }

    def root_artifact_id(record: Dict[str, Any]) -> str:
        current = str(record.get("artifact_id") or "").strip()
        seen: set[str] = set()
        while current and current not in seen:
            seen.add(current)
            parent = str((artifact_lookup.get(current) or {}).get("parent_artifact_id") or "").strip()
            if not parent:
                break
            current = parent
        return current or str(record.get("artifact_id") or "").strip()

    preferred_by_root: Dict[str, tuple[int, Dict[str, Any]]] = {}
    for index, artifact in enumerate(candidate_artifacts):
        kind = str(artifact.get("kind") or "").strip()
        root_id = root_artifact_id(artifact)
        existing = preferred_by_root.get(root_id)
        is_narrowed = kind in {"domestic_narrowed", "macro_narrowed"}
        score = 2 if is_narrowed else 1
        if existing is None:
            preferred_by_root[root_id] = (index, artifact)
            continue
        existing_index, existing_artifact = existing
        existing_kind = str(existing_artifact.get("kind") or "").strip()
        existing_score = 2 if existing_kind in {"domestic_narrowed", "macro_narrowed"} else 1
        if score > existing_score or (score == existing_score and index > existing_index):
            preferred_by_root[root_id] = (index, artifact)

    run_artifacts = [artifact for _, artifact in sorted(preferred_by_root.values(), key=lambda item: item[0])]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Summary"])
    sheet.append(["Question", user_message])
    sheet.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")])

    source_lines: List[str] = []
    seen_sources: set[str] = set()
    for artifact in run_artifacts:
        refs = artifact.get("source_references") if isinstance(artifact.get("source_references"), list) else []
        for item in refs:
            if not isinstance(item, dict):
                continue
            line = " | ".join(
                part
                for part in [
                    str(item.get("provider") or "").strip(),
                    str(item.get("dataset_id") or item.get("series_id") or "").strip(),
                    str(item.get("title") or item.get("indicator") or "").strip(),
                    str(item.get("url") or "").strip(),
                ]
                if part
            ).strip()
            if line and line not in seen_sources:
                seen_sources.add(line)
                source_lines.append(line)

    if source_lines:
        sheet.append(["Sources", _safe_sheet_text(source_lines[0])])
        for line in source_lines[1:12]:
            sheet.append(["", _safe_sheet_text(line)])

    sheet.append([])

    chart_headers, chart_rows = _chart_table(chart_spec or {})
    if chart_headers and chart_rows:
        sheet.append(["Presented data"])
        _write_table(sheet, chart_headers, chart_rows)
        sheet.append([])

    if run_artifacts:
        sheet.append(["Retrieved data"])
        _write_table(
            sheet,
            ["Artifact", "Kind", "Summary"],
            [
                [
                    str(item.get("label") or item.get("artifact_id") or "").strip(),
                    str(item.get("kind") or "").strip(),
                    str(item.get("summary") or "").strip(),
                ]
                for item in run_artifacts
            ],
        )

    widths = {"A": 18, "B": 44, "C": 24, "D": 24, "E": 20, "F": 20}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    used_sheet_names = {"Summary"}
    for artifact in run_artifacts:
        raw_sheet = workbook.create_sheet(
            title=_safe_sheet_name(
                str(artifact.get("label") or artifact.get("artifact_id") or "Raw data"),
                used_sheet_names,
            )
        )
        _write_raw_artifact_sheet(raw_sheet, artifact)

    _apply_export_theme(workbook)
    run_dir = _ensure_runtime_dirs(conversation_id)
    download_filename = _safe_export_filename(user_message)
    path = run_dir / "artifacts" / f"answer_export_{len(state.artifacts) + 1:03d}.xlsx"
    workbook.save(path)
    source_references: List[Dict[str, Any]] = []
    for artifact in run_artifacts:
        refs = artifact.get("source_references") if isinstance(artifact.get("source_references"), list) else []
        source_references.extend([item for item in refs if isinstance(item, dict)])
    record = _make_artifact_record(
        state=state,
        path=path,
        kind="answer_export",
        label="Excel export",
        summary=_truncate(f"Summary workbook with retrieved data for '{user_message}'.", 300),
        source_references=source_references[:12],
    )
    record["download_filename"] = download_filename
    state.latest_export_artifact_id = record["artifact_id"]
    return record["artifact_id"]


async def _generate_response_async(
    conversation_id: str,
    user_input: str,
    store: ConversationStore,
    status_callback: Callable[[str], None],
    user_id: str = "",
    project_id: str = "",
    project_name: str = "",
) -> str:
    set_default_openai_key(settings.openai_api_key, use_for_tracing=False)
    cancel_event = _acquire_cancellation_event(conversation_id)
    state = store.load(conversation_id)
    resolved_user_id = user_id or str(getattr(state, "user_id", "") or "")
    resolved_project_id = project_id or str(getattr(state, "project_id", "") or "")
    resolved_project_name = project_name or str(getattr(state, "project_name", "") or "")
    if resolved_user_id and resolved_project_id:
        state.messages = fetch_project_chat_messages(
            user_id=resolved_user_id,
            project_id=resolved_project_id,
        )
    run_dir = _ensure_runtime_dirs(conversation_id)
    run_artifact_start_index = len(state.artifacts)
    processed_tool_output_call_ids: set[str] = set()
    tool_call_names: Dict[str, str] = {}
    tool_call_args_by_id: Dict[str, Dict[str, Any]] = {}
    tool_call_started_at: Dict[str, float] = {}
    last_status = ""
    saved_progress_messages: List[str] = []
    _append_trace_event(
        conversation_id,
        {
            "event": "run_started",
            "user_input": _truncate(user_input, 400),
            "model": settings.openai_model,
            "reasoning_effort": settings.openai_reasoning_effort,
        },
    )

    def emit_status(message: str) -> None:
        nonlocal last_status
        normalized = str(message or "").strip()
        if not normalized or normalized == last_status:
            return
        last_status = normalized
        if not saved_progress_messages or saved_progress_messages[-1] != normalized:
            saved_progress_messages.append(normalized)
        status_callback(normalized)

    state.messages.append({"role": "user", "content": user_input})
    state.latest_export_artifact_id = ""
    state.latest_export_request = None
    state.latest_export_status = ""
    if user_id:
        state.user_id = user_id
    if project_id:
        state.project_id = project_id
    if project_name:
        state.project_name = project_name
    store.save(state)
    session = NisabaProjectSession(
        conversation_id,
        _agent_session_items_from_chat_history(state.messages),
    )
    project_compact_memory = fetch_project_compact_memory(
        user_id=resolved_user_id,
        project_id=resolved_project_id,
    )
    model_builder_state = fetch_model_builder_state(
        user_id=resolved_user_id,
        project_id=resolved_project_id,
    )

    code_container_id = _create_code_container(conversation_id)
    _append_trace_event(
        conversation_id,
        {
            "event": "code_container_created",
            "container_id": code_container_id,
        },
    )

    agent = _build_agent(code_container_id)
    integrated_server = _integrated_mcp_server_for_conversation(conversation_id, code_container_id)
    agent.mcp_servers = [integrated_server]
    runtime_context = AgentRuntimeContext(
        conversation_id=conversation_id,
        store=store,
        code_container_id=code_container_id,
        status_callback=emit_status,
        user_id=resolved_user_id,
        project_id=resolved_project_id,
        project_name=resolved_project_name,
        project_compact_memory=project_compact_memory,
    )

    _ensure_not_cancelled(conversation_id, cancel_event, "before_run")

    try:
        async with integrated_server:
            result = await Runner.run(
                agent,
                input=_build_agent_input(
                    user_input,
                    project_compact_memory,
                    model_builder_state,
                    state.pending_validated_variable_candidate,
                ),
                context=runtime_context,
                session=session,
                max_turns=30,
            )

            for item in list(getattr(result, "new_items", []) or []):
                _ensure_not_cancelled(conversation_id, cancel_event, "after_run_item")
                item_type = str(getattr(item, "type", "") or "").strip()
                if item_type in {"tool_search_call_item", "tool_search_output_item"}:
                    payload_preview = _event_payload_preview(_extract_raw_item(item))
                    event_name = "tool_search_called" if item_type == "tool_search_call_item" else "tool_search_output_created"
                    logger.info(
                        "Tool search event cid=%s event=%s payload=%s",
                        conversation_id,
                        event_name,
                        payload_preview or "-",
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": event_name,
                            "payload_preview": payload_preview,
                        },
                    )
                    continue

                if item_type == "tool_call_item":
                    tool_name = _extract_tool_name(item)
                    call_id = _extract_call_id(item)
                    tool_args = _extract_tool_arguments(item)
                    raw_item_type = _extract_raw_item_type(item)
                    server_label = _extract_mcp_server_label(item)
                    transport = _tool_transport(tool_name, item)
                    if call_id and tool_name:
                        tool_call_names[call_id] = tool_name
                        tool_call_args_by_id[call_id] = tool_args
                        tool_call_started_at[call_id] = time.perf_counter()
                    logger.info(
                        "Tool call start cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s args=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        json.dumps(
                            _display_tool_args_summary(state, tool_name, tool_args),
                            ensure_ascii=False,
                            sort_keys=True,
                        ),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_called",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "args": tool_args,
                            "args_summary": _display_tool_args_summary(state, tool_name, tool_args),
                            "raw_item_preview": _event_payload_preview(_extract_raw_item(item)),
                        },
                    )
                    continue

                if item_type != "tool_call_output_item":
                    continue

                call_id = _extract_call_id(item)
                if call_id and call_id in processed_tool_output_call_ids:
                    continue

                tool_name = tool_call_names.get(call_id or "", "") or _extract_tool_name(item)
                output_payload = _extract_tool_output_payload(item)
                raw_item_type = _extract_raw_item_type(item)
                server_label = _extract_mcp_server_label(item)
                transport = _tool_transport(tool_name, item)
                duration_ms = None
                if call_id:
                    started_at = tool_call_started_at.pop(call_id, None)
                    if started_at is not None:
                        duration_ms = int((time.perf_counter() - started_at) * 1000)
                if isinstance(output_payload, dict):
                    logger.info(
                        "Tool call success cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s duration_ms=%s summary=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        duration_ms if duration_ms is not None else -1,
                        json.dumps(_tool_output_summary(output_payload), ensure_ascii=False, sort_keys=True),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_output",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "duration_ms": duration_ms,
                            "output_summary": _tool_output_summary(output_payload),
                            "output_preview": _event_payload_preview(output_payload),
                        },
                    )
                    record = _persist_retrieval_artifact(
                        state=state,
                        conversation_id=conversation_id,
                        payload=output_payload,
                    )
                    if record:
                        call_args = tool_call_args_by_id.get(call_id or "", {})
                        if str(tool_name or "").strip() in {"retrieve", "narrow_artifact"} and (
                            call_id or tool_name or call_args
                        ):
                            record["tool_call"] = {
                                "call_id": call_id or "",
                                "tool_name": tool_name or "",
                                "args": call_args if isinstance(call_args, dict) else {},
                            }
                        store.save(state)
                        _append_trace_event(
                            conversation_id,
                            {
                                "event": "artifact_registered",
                                "call_id": call_id or "",
                                "tool_name": tool_name or "",
                                "artifact_id": str(record.get("artifact_id") or ""),
                                "artifact_kind": str(record.get("kind") or ""),
                                "artifact_label": str(record.get("label") or ""),
                                "artifact_path": str(record.get("path") or ""),
                            },
                        )
                else:
                    logger.info(
                        "Tool call complete cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s duration_ms=%s summary=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        duration_ms if duration_ms is not None else -1,
                        json.dumps(_tool_output_summary(output_payload), ensure_ascii=False, sort_keys=True),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_output",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "duration_ms": duration_ms,
                            "output_summary": _tool_output_summary(output_payload),
                            "output_preview": _event_payload_preview(getattr(item, "output", None) or _extract_raw_item(item)),
                        },
                    )
                if call_id:
                    processed_tool_output_call_ids.add(call_id)

            _ensure_not_cancelled(conversation_id, cancel_event, "after_run")
            final_answer = str(result.final_output or "").strip()
            usage = getattr(getattr(result, "context_wrapper", None), "usage", None)
    finally:
        _release_cancellation_event(conversation_id)

    if not final_answer:
        if saved_progress_messages:
            final_answer = (
                "I completed part of the request, but I could not finish the final synthesis cleanly. "
                "What I completed: " + "; ".join(saved_progress_messages[-5:])
            )
        else:
            final_answer = "I started the request but could not complete a reliable final answer from the available run state."

    _append_trace_event(
        conversation_id,
        {
            "event": "run_completed",
            "final_answer_preview": _truncate(final_answer, 600),
        },
    )

    usage_input_tokens = _safe_int(getattr(usage, "input_tokens", 0))
    usage_output_tokens = _safe_int(getattr(usage, "output_tokens", 0))
    usage_input_details = getattr(usage, "input_tokens_details", None)
    usage_cached_input_tokens = _safe_int(getattr(usage_input_details, "cached_tokens", 0))
    run_cost = _build_run_cost_payload(
        input_tokens=usage_input_tokens,
        cached_input_tokens=usage_cached_input_tokens,
        output_tokens=usage_output_tokens,
        model=settings.openai_model,
    )

    state = store.load(conversation_id)
    current_run_start = state.active_run_message_count if isinstance(state.active_run_message_count, int) else 0
    existing_current_progress = [
        str(message.get("content") or "").strip()
        for message in list(state.messages or [])[current_run_start:]
        if isinstance(message, dict) and str(message.get("role") or "").strip() == "progress"
    ]
    for progress_message in saved_progress_messages:
        if progress_message not in existing_current_progress:
            state.messages.append({"role": "progress", "content": progress_message})
            existing_current_progress.append(progress_message)
    state.messages.append({"role": "assistant", "content": final_answer, "run_cost": run_cost})
    persisted_chat = persist_project_chat_run(
        user_id=resolved_user_id,
        project_id=resolved_project_id,
        conversation_id=conversation_id,
        user_message=user_input,
        progress_notes=saved_progress_messages,
        final_response=final_answer,
        run_cost=run_cost,
    )
    if persisted_chat:
        _append_trace_event(
            conversation_id,
            {
                "event": "supabase_chat_history_persisted",
                "project_id": resolved_project_id,
            },
        )
    has_exportable_artifacts = len(state.artifacts) > run_artifact_start_index
    has_chart = _parse_chart_spec_from_markdown(final_answer) is not None
    if has_exportable_artifacts or has_chart:
        state.latest_export_status = "processing"
        state.latest_export_request = {
            "user_message": user_input,
            "final_answer": final_answer,
            "run_artifact_start_index": run_artifact_start_index,
        }
    else:
        state.latest_export_status = ""
        state.latest_export_request = None
    store.save(state)
    _schedule_project_memory_compaction(
        user_id=resolved_user_id,
        project_id=resolved_project_id,
        project_name=resolved_project_name,
        conversation_id=conversation_id,
        messages=list(state.messages),
    )
    return final_answer


def generate_response(
    conversation_id: str,
    user_input: str,
    store: ConversationStore,
    status_callback: Callable[[str], None],
    user_id: str = "",
    project_id: str = "",
    project_name: str = "",
) -> str:
    return asyncio.run(
        _generate_response_async(
            conversation_id=conversation_id,
            user_input=user_input,
            store=store,
            status_callback=status_callback,
            user_id=user_id,
            project_id=project_id,
            project_name=project_name,
        )
    )
